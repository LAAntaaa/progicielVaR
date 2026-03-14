"""
╔══════════════════════════════════════════════════════════════╗
║        VaR ANALYTICS SUITE  ·  Streamlit App v4.0           ║
║        Département Gestion des Risques de Marché            ╠
║        Améliorations v4.0 :                                  ║
║          · Optimisation de portefeuille (Markowitz)          ║
║          · Stress-Testing (scénarios historiques)            ║
║          · ES analytique TVE-GARCH                           ║
║          · Matrice de corrélation interactive                ║
╚══════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
from scipy.optimize import minimize, differential_evolution
import io, os, warnings
warnings.filterwarnings("ignore")

# ── Imports optionnels ────────────────────────────────────────────────────────
try:
    import yfinance as yf
    HAS_YF = True
except ImportError:
    HAS_YF = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable, PageBreak, Image)
    from reportlab.lib.colors import HexColor
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG STREAMLIT + CSS
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="VaR Analytics Suite",
    page_icon="📉",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* ══════════════════════════════════════════════════════════
   VaR Analytics Suite — Design System "Obsidian Terminal"
   Fonts : Outfit (display) + DM Mono (data)
   Palette : #07090F base · #00D4FF cyan · #F0B429 amber
             #FF4D6D risk-red · #00C896 signal-green
   ══════════════════════════════════════════════════════════ */

@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=DM+Mono:ital,wght@0,300;0,400;0,500;1,300&display=swap');

/* ── Tokens ───────────────────────────────────────────── */
:root {
  --bg-void:       #07090F;
  --bg-deep:       #0C0F1A;
  --bg-surface:    #111827;
  --bg-raised:     #1a2235;
  --bg-glass:      rgba(17,24,37,0.72);

  --cyan:          #00D4FF;
  --cyan-dim:      rgba(0,212,255,0.12);
  --cyan-glow:     rgba(0,212,255,0.25);

  --amber:         #F0B429;
  --amber-dim:     rgba(240,180,41,0.10);
  --amber-glow:    rgba(240,180,41,0.22);

  --risk-red:      #FF4D6D;
  --risk-red-dim:  rgba(255,77,109,0.10);

  --signal-green:  #00C896;
  --signal-dim:    rgba(0,200,150,0.10);

  --txt-primary:   #E8EDF5;
  --txt-secondary: #7A8BA8;
  --txt-muted:     #3D4F6B;

  --border-subtle: rgba(0,212,255,0.08);
  --border-dim:    rgba(0,212,255,0.16);

  --radius-sm:     6px;
  --radius-md:     10px;
  --radius-lg:     16px;

  --shadow-card:   0 4px 24px rgba(0,0,0,0.5), 0 1px 0 rgba(255,255,255,0.03) inset;
  --shadow-glow-c: 0 0 20px rgba(0,212,255,0.15);
  --shadow-glow-a: 0 0 20px rgba(240,180,41,0.15);

  --font-ui:   'Outfit', sans-serif;
  --font-data: 'DM Mono', monospace;
}

/* ── Reset & Base ────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: var(--font-ui) !important;
    -webkit-font-smoothing: antialiased;
    text-rendering: optimizeLegibility;
}

/* ── Fond animé : halos en mouvement + grille pulsante ────── */
.stApp {
    background-color: var(--bg-void) !important;
    background-image: none !important;
    position: relative !important;
    overflow-x: hidden !important;
}

/* Halo cyan animé (coin haut-gauche → centre) */
.stApp::before {
    content: '' !important;
    position: fixed !important;
    top: -20% !important;
    left: -10% !important;
    width: 60vw !important;
    height: 60vw !important;
    border-radius: 50% !important;
    background: radial-gradient(circle, rgba(0,212,255,0.10) 0%, transparent 65%) !important;
    animation: haloMoveCyan 22s ease-in-out infinite alternate !important;
    pointer-events: none !important;
    z-index: 0 !important;
    filter: blur(60px) !important;
}

/* Halo amber animé (coin bas-droit → centre) */
.stApp::after {
    content: '' !important;
    position: fixed !important;
    bottom: -15% !important;
    right: -8% !important;
    width: 50vw !important;
    height: 50vw !important;
    border-radius: 50% !important;
    background: radial-gradient(circle, rgba(240,180,41,0.08) 0%, transparent 65%) !important;
    animation: haloMoveAmber 28s ease-in-out infinite alternate !important;
    pointer-events: none !important;
    z-index: 0 !important;
    filter: blur(70px) !important;
}

@keyframes haloMoveCyan {
    0%   { top: -20%; left: -10%; opacity: 0.7; }
    30%  { top:  10%; left:  15%; opacity: 1.0; }
    60%  { top: -5%;  left:  35%; opacity: 0.8; }
    100% { top:  5%;  left: -5%;  opacity: 0.9; }
}
@keyframes haloMoveAmber {
    0%   { bottom: -15%; right: -8%;  opacity: 0.6; }
    35%  { bottom:  20%; right: 20%;  opacity: 1.0; }
    70%  { bottom:  5%;  right: 40%;  opacity: 0.7; }
    100% { bottom: -5%;  right: 10%;  opacity: 0.9; }
}

/* Grille de points pulsante — sur stAppViewContainer */
[data-testid="stAppViewContainer"] {
    position: relative !important;
}
[data-testid="stAppViewContainer"]::before {
    content: '' !important;
    position: fixed !important;
    inset: 0 !important;
    background-image: radial-gradient(circle, rgba(0,212,255,0.14) 1px, transparent 1px) !important;
    background-size: 52px 52px !important;
    animation: gridPulse 8s ease-in-out infinite !important;
    pointer-events: none !important;
    z-index: 0 !important;
}
@keyframes gridPulse {
    0%,100% { opacity: 0.15; }
    50%     { opacity: 0.28; }
}

/* Grain de bruit statique */
[data-testid="stAppViewContainer"]::after {
    content: '' !important;
    position: fixed !important;
    inset: 0 !important;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='200' height='200'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.8' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='200' height='200' filter='url(%23n)' opacity='0.022'/%3E%3C/svg%3E") !important;
    pointer-events: none !important;
    z-index: 0 !important;
}

/* Contenu au-dessus des effets de fond */
[data-testid="stAppViewContainer"] > section.main {
    position: relative !important;
    z-index: 1 !important;
}


/* ── Sidebar ─────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background: var(--bg-deep) !important;
    border-right: 1px solid var(--border-subtle) !important;
}
[data-testid="stSidebar"]::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, var(--cyan), var(--amber), var(--cyan));
    background-size: 200% 100%;
    animation: shimmer 4s linear infinite;
}
@keyframes shimmer {
    0%   { background-position: 200% 0; }
    100% { background-position: -200% 0; }
}
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
    color: var(--txt-secondary) !important;
    font-size: 12px !important;
    font-family: var(--font-ui) !important;
}

/* ── Titres ──────────────────────────────────────────────── */
h1 {
    font-family: var(--font-ui) !important;
    font-size: 2rem !important;
    font-weight: 800 !important;
    letter-spacing: -0.5px !important;
    background: linear-gradient(120deg, #fff 0%, var(--cyan) 45%, var(--amber) 100%) !important;
    -webkit-background-clip: text !important;
    -webkit-text-fill-color: transparent !important;
    background-clip: text !important;
    line-height: 1.15 !important;
    margin-bottom: 4px !important;
}
h2 {
    color: var(--txt-primary) !important;
    font-size: 1.1rem !important;
    font-weight: 600 !important;
    letter-spacing: -0.2px !important;
}
h3 {
    color: var(--amber) !important;
    font-size: 0.95rem !important;
    font-weight: 600 !important;
}

/* ── Metric cards — glassmorphisme ───────────────────────── */
[data-testid="metric-container"] {
    background: var(--bg-glass) !important;
    border: 1px solid var(--border-dim) !important;
    border-radius: var(--radius-md) !important;
    padding: 16px 18px 14px !important;
    box-shadow: var(--shadow-card) !important;
    backdrop-filter: blur(12px) !important;
    -webkit-backdrop-filter: blur(12px) !important;
    position: relative; overflow: hidden;
    transition: border-color 0.25s, box-shadow 0.25s !important;
}
[data-testid="metric-container"]:hover {
    border-color: rgba(0,212,255,0.35) !important;
    box-shadow: var(--shadow-card), var(--shadow-glow-c) !important;
}
[data-testid="metric-container"]::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg,
        transparent 0%, var(--cyan) 30%, var(--amber) 70%, transparent 100%);
    opacity: 0.6;
}
[data-testid="metric-container"]::after {
    content: '';
    position: absolute; bottom: 0; right: 0;
    width: 60px; height: 60px;
    background: radial-gradient(circle, var(--cyan-dim) 0%, transparent 70%);
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    color: var(--txt-secondary) !important;
    font-size: 9.5px !important;
    font-family: var(--font-data) !important;
    text-transform: uppercase !important;
    letter-spacing: 1.2px !important;
    font-weight: 400 !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: var(--cyan) !important;
    font-size: 1.55rem !important;
    font-family: var(--font-data) !important;
    font-weight: 500 !important;
    letter-spacing: -0.5px !important;
}
[data-testid="metric-container"] [data-testid="stMetricDelta"] {
    font-family: var(--font-data) !important;
    font-size: 10px !important;
}

/* ── Boutons ─────────────────────────────────────────────── */
.stButton > button {
    background: var(--bg-raised) !important;
    color: var(--txt-primary) !important;
    border: 1px solid var(--border-dim) !important;
    border-radius: var(--radius-sm) !important;
    font-family: var(--font-ui) !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    padding: 9px 20px !important;
    letter-spacing: 0.2px !important;
    transition: all 0.18s cubic-bezier(.4,0,.2,1) !important;
    position: relative !important;
    overflow: hidden !important;
}
.stButton > button::before {
    content: '';
    position: absolute; inset: 0;
    background: linear-gradient(135deg, var(--cyan-dim), transparent);
    opacity: 0;
    transition: opacity 0.18s;
}
.stButton > button:hover {
    border-color: var(--cyan) !important;
    color: var(--cyan) !important;
    box-shadow: 0 0 0 1px var(--cyan-dim), var(--shadow-glow-c) !important;
    transform: translateY(-1px) !important;
}
.stButton > button:hover::before { opacity: 1 !important; }
.stButton > button:active { transform: translateY(0) !important; }

/* Bouton primary (CTA principal) */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #0099bb 0%, var(--cyan) 50%, #00aacc 100%) !important;
    color: var(--bg-void) !important;
    border: none !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 20px var(--cyan-glow) !important;
}
.stButton > button[kind="primary"]:hover {
    box-shadow: 0 6px 28px rgba(0,212,255,0.45) !important;
    transform: translateY(-2px) !important;
    color: var(--bg-void) !important;
}

/* ── Selectbox / Multiselect ─────────────────────────────── */
.stSelectbox > div > div,
.stMultiSelect > div > div {
    background: var(--bg-raised) !important;
    border: 1px solid var(--border-dim) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--txt-primary) !important;
    font-family: var(--font-ui) !important;
    transition: border-color 0.15s !important;
}
.stSelectbox > div > div:focus-within,
.stMultiSelect > div > div:focus-within {
    border-color: var(--cyan) !important;
    box-shadow: 0 0 0 3px var(--cyan-dim) !important;
}

/* ── Number input ────────────────────────────────────────── */
.stNumberInput > div > div > input {
    background: var(--bg-raised) !important;
    border: 1px solid var(--border-dim) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--txt-primary) !important;
    font-family: var(--font-data) !important;
}
.stNumberInput > div > div > input:focus {
    border-color: var(--cyan) !important;
    box-shadow: 0 0 0 3px var(--cyan-dim) !important;
}

/* ── Radio buttons ───────────────────────────────────────── */
[data-testid="stRadio"] > div { gap: 8px !important; }
[data-testid="stRadio"] label {
    background: var(--bg-raised) !important;
    border: 1px solid var(--border-subtle) !important;
    border-radius: var(--radius-sm) !important;
    padding: 6px 14px !important;
    color: var(--txt-secondary) !important;
    font-size: 12px !important;
    cursor: pointer;
    transition: all 0.15s !important;
}
[data-testid="stRadio"] label:has(input:checked) {
    border-color: var(--cyan) !important;
    color: var(--cyan) !important;
    background: var(--cyan-dim) !important;
}

/* ── Slider ──────────────────────────────────────────────── */
.stSlider [data-baseweb="slider"] [role="slider"] {
    background: var(--cyan) !important;
    box-shadow: 0 0 8px var(--cyan-glow) !important;
    border: 2px solid var(--bg-void) !important;
}
.stSlider [data-baseweb="slider"] [data-testid="stTickBar"] > div {
    background: linear-gradient(90deg, var(--cyan), var(--amber)) !important;
}

/* ── Date input ──────────────────────────────────────────── */
.stDateInput > div > div > input {
    background: var(--bg-raised) !important;
    border: 1px solid var(--border-dim) !important;
    color: var(--txt-primary) !important;
    border-radius: var(--radius-sm) !important;
    font-family: var(--font-data) !important;
}

/* ── DataFrames ──────────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border: 1px solid var(--border-dim) !important;
    border-radius: var(--radius-md) !important;
    overflow: hidden !important;
    box-shadow: var(--shadow-card) !important;
}
[data-testid="stDataFrame"] th {
    background: var(--bg-surface) !important;
    color: var(--txt-secondary) !important;
    font-family: var(--font-data) !important;
    font-size: 10.5px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.8px !important;
    border-bottom: 1px solid var(--border-dim) !important;
}
[data-testid="stDataFrame"] td {
    font-family: var(--font-data) !important;
    font-size: 12px !important;
    color: var(--txt-primary) !important;
}

/* ── Expander ────────────────────────────────────────────── */
.streamlit-expanderHeader {
    background: var(--bg-surface) !important;
    border: 1px solid var(--border-subtle) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--txt-secondary) !important;
    font-family: var(--font-ui) !important;
    font-size: 13px !important;
    transition: all 0.15s !important;
}
.streamlit-expanderHeader:hover {
    border-color: var(--border-dim) !important;
    color: var(--txt-primary) !important;
}
.streamlit-expanderContent {
    border: 1px solid var(--border-subtle) !important;
    border-top: none !important;
    border-radius: 0 0 var(--radius-sm) var(--radius-sm) !important;
    background: var(--bg-deep) !important;
}

/* ── Alertes Streamlit ───────────────────────────────────── */
.stAlert {
    border-radius: var(--radius-sm) !important;
    font-family: var(--font-ui) !important;
    font-size: 13px !important;
}

/* ── Divider ─────────────────────────────────────────────── */
hr {
    border: none !important;
    border-top: 1px solid var(--border-subtle) !important;
    margin: 20px 0 !important;
}

/* ═══════════════════════════════════════════════════════════
   COMPOSANTS CUSTOM
   ═══════════════════════════════════════════════════════════ */

/* ── VaR card ────────────────────────────────────────────── */
.var-card {
    background: var(--bg-glass);
    backdrop-filter: blur(10px);
    -webkit-backdrop-filter: blur(10px);
    border: 1px solid var(--border-dim);
    border-radius: var(--radius-md);
    padding: 18px 20px 16px;
    margin-bottom: 12px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s, transform 0.2s, box-shadow 0.2s;
    cursor: default;
}
.var-card:hover {
    border-color: rgba(0,212,255,0.32);
    transform: translateY(-2px);
    box-shadow: var(--shadow-card), var(--shadow-glow-c);
}
.var-card::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg,
        transparent, var(--cyan) 30%, var(--amber) 70%, transparent);
    opacity: 0.5;
}
.var-card::after {
    content: '';
    position: absolute; top: -30px; right: -30px;
    width: 90px; height: 90px;
    background: radial-gradient(circle, var(--cyan-dim) 0%, transparent 65%);
    pointer-events: none;
}
.var-card-title {
    font-size: 10px;
    color: var(--txt-muted);
    text-transform: uppercase;
    letter-spacing: 1.8px;
    font-family: var(--font-data);
    font-weight: 400;
    margin-bottom: 8px;
}
.var-card-value {
    font-size: 26px;
    font-weight: 300;
    color: var(--risk-red);
    font-family: var(--font-data);
    letter-spacing: -1px;
    line-height: 1.1;
}
.var-card-pct {
    font-size: 11px;
    color: var(--txt-secondary);
    font-family: var(--font-data);
    margin-top: 3px;
}
.var-card-es {
    font-size: 12px;
    color: rgba(255,77,109,0.7);
    font-family: var(--font-data);
    margin-top: 6px;
    padding-top: 6px;
    border-top: 1px solid var(--risk-red-dim);
}

/* ── Badge ───────────────────────────────────────────────── */
.badge-rec {
    display: inline-flex;
    align-items: center;
    gap: 3px;
    background: linear-gradient(135deg, var(--cyan), #0099bb);
    color: var(--bg-void);
    font-size: 8px;
    font-weight: 700;
    padding: 2px 8px 2px 6px;
    border-radius: 20px;
    letter-spacing: 0.8px;
    font-family: var(--font-data);
    margin-left: 8px;
    text-transform: uppercase;
    vertical-align: middle;
    box-shadow: 0 2px 8px var(--cyan-glow);
}

/* ── Section header ──────────────────────────────────────── */
.section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 28px 0 14px;
    font-size: 11px;
    font-weight: 600;
    color: var(--txt-secondary);
    text-transform: uppercase;
    letter-spacing: 1.5px;
    font-family: var(--font-data);
}
.section-header::before {
    content: '';
    display: inline-block;
    width: 3px; height: 16px;
    background: linear-gradient(180deg, var(--cyan), var(--amber));
    border-radius: 2px;
    flex-shrink: 0;
}
.section-header::after {
    content: '';
    flex: 1;
    height: 1px;
    background: linear-gradient(90deg, var(--border-dim), transparent);
}

/* ── Info box ────────────────────────────────────────────── */
.info-box {
    background: rgba(0,212,255,0.04);
    border: 1px solid rgba(0,212,255,0.14);
    border-left: 3px solid var(--cyan);
    border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
    padding: 14px 18px;
    margin: 14px 0;
    font-size: 12.5px;
    color: var(--txt-secondary);
    line-height: 1.75;
    font-family: var(--font-ui);
}

/* ── Stress card ─────────────────────────────────────────── */
.stress-card {
    background: var(--bg-glass);
    backdrop-filter: blur(8px);
    border: 1px solid rgba(255,77,109,0.18);
    border-radius: var(--radius-md);
    padding: 16px 18px;
    margin-bottom: 10px;
    position: relative;
    overflow: hidden;
    transition: transform 0.18s, border-color 0.18s;
}
.stress-card:hover {
    border-color: rgba(255,77,109,0.38);
    transform: translateY(-1px);
}
.stress-card::after {
    content: '';
    position: absolute; top: -20px; right: -20px;
    width: 70px; height: 70px;
    background: radial-gradient(circle, var(--risk-red-dim), transparent 65%);
}
.stress-title {
    font-size: 9.5px;
    color: rgba(255,77,109,0.7);
    font-family: var(--font-data);
    text-transform: uppercase;
    letter-spacing: 1.5px;
    margin-bottom: 8px;
}
.stress-val {
    font-size: 24px;
    font-weight: 300;
    color: var(--risk-red);
    font-family: var(--font-data);
    letter-spacing: -0.8px;
}

/* ── Markowitz card ──────────────────────────────────────── */
.markowitz-card {
    background: rgba(0,200,150,0.04);
    border: 1px solid rgba(0,200,150,0.18);
    border-radius: var(--radius-md);
    padding: 16px 18px;
    margin-bottom: 10px;
    transition: transform 0.18s, border-color 0.18s;
}
.markowitz-card:hover {
    border-color: rgba(0,200,150,0.35);
    transform: translateY(-1px);
}

/* ── Spinner / Progress ──────────────────────────────────── */
[data-testid="stSpinner"] {
    color: var(--cyan) !important;
}

/* ── Scrollbar custom ────────────────────────────────────── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: var(--bg-deep); }
::-webkit-scrollbar-thumb {
    background: var(--border-dim);
    border-radius: 10px;
}
::-webkit-scrollbar-thumb:hover { background: var(--cyan); }

/* ── Animations d'entrée ─────────────────────────────────── */
@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(12px); }
    to   { opacity: 1; transform: translateY(0); }
}
.var-card, .stress-card, .markowitz-card, [data-testid="metric-container"] {
    animation: fadeSlideUp 0.35s cubic-bezier(.4,0,.2,1) both;
}

/* ── Suppression éléments Streamlit indésirables ─────────── */
#MainMenu, footer, [data-testid="stToolbar"] { display: none !important; }
[data-testid="stHeader"] { background: transparent !important; }

</style>
""", unsafe_allow_html=True)

# ── Sidebar toggle : JS MutationObserver (robuste toutes versions Streamlit) ──
st.markdown("""
<script>
(function() {
  // Styles à appliquer sur le bouton collapse/expand
  const BTN_STYLE = {
    opacity: '1',
    visibility: 'visible',
    background: '#111827',
    border: '1px solid rgba(0,212,255,0.40)',
    borderLeft: 'none',
    borderRadius: '0 8px 8px 0',
    boxShadow: '3px 0 20px rgba(0,212,255,0.15)',
    transition: 'all 0.2s ease',
    zIndex: '99999',
    width: '28px',
    minWidth: '28px',
    height: '52px',
    cursor: 'pointer',
    display: 'flex',
    alignItems: 'center',
    justifyContent: 'center',
    position: 'fixed',
  };
  const SVG_STYLE = {
    color: '#00D4FF',
    fill: '#00D4FF',
    stroke: '#00D4FF',
    width: '16px',
    height: '16px',
  };

  function styleBtn(btn) {
    if (!btn) return;
    Object.assign(btn.style, BTN_STYLE);
    btn.onmouseenter = () => {
      btn.style.background = '#1a2235';
      btn.style.borderColor = '#00D4FF';
      btn.style.boxShadow = '3px 0 28px rgba(0,212,255,0.30)';
    };
    btn.onmouseleave = () => {
      btn.style.background = '#111827';
      btn.style.borderColor = 'rgba(0,212,255,0.40)';
      btn.style.boxShadow = '3px 0 20px rgba(0,212,255,0.15)';
    };
    // Styler les SVG à l'intérieur
    btn.querySelectorAll('svg').forEach(svg => Object.assign(svg.style, SVG_STYLE));
    btn.querySelectorAll('button').forEach(b => {
      b.style.background = 'transparent';
      b.style.border = 'none';
      b.style.width = '100%';
      b.style.height = '100%';
      b.style.color = '#00D4FF';
      b.style.opacity = '1';
    });
  }

  // Sélecteurs connus selon la version de Streamlit
  const SELECTORS = [
    '[data-testid="stSidebarCollapsedControl"]',
    '[data-testid="collapsedControl"]',
    '.st-emotion-cache-1rtdyuf',   // classe générée fréquente
    '.st-emotion-cache-czk5ss',
  ];

  function applyToAll() {
    SELECTORS.forEach(sel => {
      document.querySelectorAll(sel).forEach(el => styleBtn(el));
    });
    // Fallback : chercher tout bouton près du bord gauche hors sidebar
    document.querySelectorAll('button').forEach(btn => {
      const rect = btn.getBoundingClientRect();
      const sidebar = document.querySelector('[data-testid="stSidebar"]');
      const isSidebarOpen = sidebar && sidebar.offsetWidth > 100;
      // Bouton dans la zone gauche, pas dans la sidebar ouverte
      if (rect.left < 40 && rect.top > 50 && !isSidebarOpen) {
        styleBtn(btn.closest('[data-testid]') || btn.parentElement || btn);
      }
    });
  }

  // Observer les mutations du DOM (Streamlit recrée les éléments à chaque rerun)
  const observer = new MutationObserver(() => applyToAll());
  observer.observe(document.body, { childList: true, subtree: true });

  // Premier appel après chargement
  if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', applyToAll);
  } else {
    applyToAll();
    setTimeout(applyToAll, 500);
    setTimeout(applyToAll, 1500);
  }
})();
</script>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# DONNÉES & CACHE
# ══════════════════════════════════════════════════════════════════════════════

ACTIFS_DISPONIBLES = {
    "Apple (AAPL)":         "AAPL",
    "Microsoft (MSFT)":     "MSFT",
    "LVMH (MC.PA)":         "MC.PA",
    "TotalEnergies (TTE)":  "TTE.PA",
    "BNP Paribas (BNP)":    "BNP.PA",
    "Nestlé (NESN.SW)":     "NESN.SW",
    "SAP (SAP)":            "SAP",
    "Airbus (AIR.PA)":      "AIR.PA",
    "Tesla (TSLA)":         "TSLA",
    "Amazon (AMZN)":        "AMZN",
    "Nvidia (NVDA)":        "NVDA",
    "Safran (SAF.PA)":      "SAF.PA",
    "L'Oréal (OR.PA)":      "OR.PA",
    "ASML (ASML.AS)":       "ASML.AS",
    "Hermès (RMS.PA)":      "RMS.PA",
}

SECTEURS = {
    "AAPL": "Technologie", "MSFT": "Technologie", "MC.PA": "Luxe",
    "TTE.PA": "Énergie",   "BNP.PA": "Finance",   "NESN.SW": "Conso.",
    "SAP": "Technologie",  "AIR.PA": "Aéronaut.", "TSLA": "Auto/Tech",
    "AMZN": "Commerce",    "NVDA": "Technologie", "SAF.PA": "Aéronaut.",
    "OR.PA": "Beauté",     "ASML.AS": "Technologie", "RMS.PA": "Luxe",
}

# Scénarios de stress historiques (chocs journaliers approximatifs)
SCENARIOS_STRESS = {
    "Crise Financière 2008 (Lehman, 15/09/2008)": {
        "description": "Faillite de Lehman Brothers — pire journée de la crise des subprimes",
        "choc_marche": -0.0469,
        "choc_vol": 3.8,
        "date": "15 sept. 2008",
        "couleur": "#e05252"
    },
    "Flash Crash 2010 (06/05/2010)": {
        "description": "Effondrement éclair de 1000 pts du DJIA en quelques minutes",
        "choc_marche": -0.034,
        "choc_vol": 2.5,
        "date": "6 mai 2010",
        "couleur": "#f97316"
    },
    "Brexit (24/06/2016)": {
        "description": "Référendum Brexit — choc sur marchés européens et GBP",
        "choc_marche": -0.0281,
        "choc_vol": 2.1,
        "date": "24 juin 2016",
        "couleur": "#a855f7"
    },
    "COVID-19 (16/03/2020)": {
        "description": "Pire séance COVID — annonce confinements mondiaux",
        "choc_marche": -0.0598,
        "choc_vol": 4.2,
        "date": "16 mars 2020",
        "couleur": "#e05252"
    },
    "Krach Obligataire 2022 (Q1)": {
        "description": "Remontée brutale des taux Fed — pire Q1 depuis 1973",
        "choc_marche": -0.0312,
        "choc_vol": 2.0,
        "date": "T1 2022",
        "couleur": "#C9A84C"
    },
    "Scénario Adverse (−3σ)": {
        "description": "Choc extrême calibré à −3 écarts-types journaliers",
        "choc_marche": None,  # calculé dynamiquement
        "choc_vol": 3.0,
        "date": "Hypothétique",
        "couleur": "#8b5cf6"
    },
}

@st.cache_data(show_spinner=False, ttl=3600)
def telecharger_donnees(tickers: list, date_debut: str, date_fin: str) -> pd.DataFrame:
    """Télécharge les cours et comble les trous par forward-fill (jours fériés)."""
    if not HAS_YF:
        return pd.DataFrame()
    try:
        data = yf.download(tickers, start=date_debut, end=date_fin,
                           auto_adjust=True, progress=False)
    except Exception:
        return pd.DataFrame()
    if data.empty:
        return pd.DataFrame()
    # Extraction colonne Close (gestion MultiIndex yfinance ≥ 0.2)
    if isinstance(data.columns, pd.MultiIndex):
        if "Close" in data.columns.get_level_values(0):
            prix = data["Close"].copy()
        else:
            prix = data.xs(data.columns.get_level_values(0)[0], axis=1, level=0)
    else:
        prix = data[["Close"]].copy() if "Close" in data.columns else data.copy()
        if len(tickers) == 1:
            prix.columns = tickers
    if isinstance(prix, pd.Series):
        prix = prix.to_frame()
    # Combler les trous dus aux jours fériés non synchronisés entre places
    prix = prix.ffill().bfill()
    # Supprimer uniquement les colonnes qui restent entièrement vides
    prix = prix.dropna(axis=1, how="all")
    return prix


def donnees_simulation(tickers: list, n_days: int = 1500) -> pd.DataFrame:
    np.random.seed(42)
    n = len(tickers)
    mu_d  = np.full(n, 0.0004)
    sig_d = np.full(n, 0.012)
    corr  = np.eye(n) + 0.4 * (np.ones((n,n)) - np.eye(n))
    L = np.linalg.cholesky(corr)
    z = np.random.randn(n_days, n) @ L.T
    lr = mu_d + sig_d * z
    mid = n_days // 2
    lr[mid:mid+20] *= 3.5
    prices = 100 * np.exp(np.cumsum(lr, axis=0))
    dates = pd.bdate_range(end="2024-12-31", periods=n_days)
    return pd.DataFrame(prices, index=dates, columns=tickers)


# ══════════════════════════════════════════════════════════════════════════════
# OPTIMISATION DE PORTEFEUILLE — MARKOWITZ
# ══════════════════════════════════════════════════════════════════════════════

def optimiser_portefeuille(rendements: pd.DataFrame, rf: float = 0.03) -> dict:
    """Calcule la frontière efficiente et 3 portefeuilles optimaux."""
    mu = rendements.mean().values * 252
    cov = rendements.cov().values * 252
    n = len(mu)
    rf_d = rf / 252

    def port_stats(w):
        ret = w @ mu
        vol = np.sqrt(w @ cov @ w)
        sharpe = (ret - rf) / vol if vol > 0 else 0
        return ret, vol, sharpe

    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1}]
    bounds = [(0.0, 1.0)] * n
    w0 = np.ones(n) / n

    # 1. Sharpe Maximum
    def neg_sharpe(w):
        r, v, _ = port_stats(w)
        return -(r - rf) / v if v > 0 else 1e10

    res_sharpe = minimize(neg_sharpe, w0, method='SLSQP',
                          bounds=bounds, constraints=constraints)
    w_sharpe = res_sharpe.x if res_sharpe.success else w0

    # 2. Variance Minimum
    def port_var(w):
        return w @ cov @ w

    res_minvar = minimize(port_var, w0, method='SLSQP',
                          bounds=bounds, constraints=constraints)
    w_minvar = res_minvar.x if res_minvar.success else w0

    # 3. Équipondéré
    w_equi = np.ones(n) / n

    # Frontière efficiente
    ret_range = np.linspace(mu.min(), mu.max(), 60)
    frontier_vols, frontier_rets = [], []
    for target_ret in ret_range:
        cons = [{"type": "eq", "fun": lambda w: np.sum(w) - 1},
                {"type": "eq", "fun": lambda w, r=target_ret: w @ mu - r}]
        res = minimize(port_var, w0, method='SLSQP', bounds=bounds, constraints=cons)
        if res.success:
            frontier_vols.append(np.sqrt(res.fun))
            frontier_rets.append(target_ret)

    tickers = list(rendements.columns)
    results = {
        "tickers": tickers,
        "mu": mu,
        "cov": cov,
        "sharpe": {
            "weights": w_sharpe,
            "stats": port_stats(w_sharpe),
            "label": "Sharpe Max."
        },
        "minvar": {
            "weights": w_minvar,
            "stats": port_stats(w_minvar),
            "label": "Variance Min."
        },
        "equi": {
            "weights": w_equi,
            "stats": port_stats(w_equi),
            "label": "Équipondéré"
        },
        "frontier": (frontier_vols, frontier_rets)
    }
    return results


def fig_frontiere_efficiente(opt_results: dict) -> plt.Figure:
    with plt.rc_context(PLT_DARK):
        fig, ax = plt.subplots(figsize=(11, 5.5))
        fig.patch.set_alpha(0)

        fv, fr = opt_results["frontier"]
        if fv:
            vols_f = [v*100 for v in fv]
            rets_f = [r*100 for r in fr]
            # Dégradé de couleur sur la frontière
            from matplotlib.collections import LineCollection
            points = np.array([vols_f, rets_f]).T.reshape(-1, 1, 2)
            segs   = np.concatenate([points[:-1], points[1:]], axis=1)
            lc = LineCollection(segs, cmap="cool", linewidth=2.8, zorder=3)
            lc.set_array(np.linspace(0, 1, len(segs)))
            ax.add_collection(lc)
            # Zone ombragée sous la frontière
            ax.fill_between(vols_f, rets_f, min(rets_f),
                            alpha=0.07, color="#00D4FF", zorder=1)

        # Actifs individuels
        mu      = opt_results["mu"]
        cov     = opt_results["cov"]
        tickers = opt_results["tickers"]
        asset_colors = ["#00D4FF","#F0B429","#00C896","#FF4D6D","#9D7FEA",
                        "#FF7849","#FFD166","#4CC9F0","#F72585","#B5E48C",
                        "#C77DFF","#06D6A0","#FFAA5B","#00B4D8","#FF9B54"]
        for i, t in enumerate(tickers):
            vol_i = np.sqrt(cov[i, i]) * 100
            mu_i  = mu[i] * 100
            col_i = asset_colors[i % len(asset_colors)]
            ax.scatter(vol_i, mu_i, s=55, color=col_i, zorder=6,
                       alpha=0.85, edgecolors="#07090F", linewidths=0.8, marker="o")
            ax.annotate(t, (vol_i, mu_i), fontsize=7.5, color=col_i,
                        fontweight="bold",
                        xytext=(5, 4), textcoords="offset points")

        # 3 portefeuilles optimaux — marqueurs matplotlib standards uniquement
        styles = [
            ("sharpe", "#C9A84C", "*",  320, "Sharpe Max."),
            ("minvar", "#1db87a", "v",  160, "Variance Min."),
            ("equi",   "#a855f7", "D",  130, "Équipondéré"),
        ]
        for key, col, marker, size, lbl in styles:
            p    = opt_results[key]
            r, v, s = p["stats"]
            # Halo lumineux derrière le marqueur
            ax.scatter(v*100, r*100, s=size*2.5, color=col,
                       alpha=0.18, zorder=7, marker=marker)
            ax.scatter(v*100, r*100, s=size, color=col, zorder=8,
                       marker=marker, edgecolors="white", linewidths=0.9,
                       label=f"{lbl}  (Sharpe {s:.2f})")
            ax.annotate(lbl, (v*100, r*100), fontsize=8.5, color=col,
                        fontweight="bold",
                        xytext=(9, 5), textcoords="offset points",
                        bbox=dict(boxstyle="round,pad=0.25", fc="#07090F",
                                  ec=col, alpha=0.75, lw=0.8))

        ax.set_xlabel("Volatilité annualisée (%)", fontsize=10, labelpad=8)
        ax.set_ylabel("Rendement annualisé (%)", fontsize=10, labelpad=8)
        ax.set_title("Frontière Efficiente de Markowitz",
                     fontsize=13, color="#F0B429", pad=12, fontweight="bold")
        legend = ax.legend(fontsize=8.5, loc="lower right",
                           framealpha=0.85, edgecolor="#1a2235",
                           facecolor="#111827", labelcolor="#d0daea")
        ax.tick_params(labelsize=8.5)
        ax.grid(True, alpha=0.18, linestyle="--")
        # Cadre sobre
        for spine in ax.spines.values():
            spine.set_edgecolor("#1a2235")
        plt.tight_layout()
        return fig


def fig_poids_portefeuilles(opt_results: dict) -> plt.Figure:
    tickers = opt_results["tickers"]
    keys    = ["sharpe", "minvar", "equi"]
    labels_ = ["Sharpe Max.", "Variance Min.", "Équipondéré"]
    accent  = ["#C9A84C", "#1db87a", "#a855f7"]

    # Palette étendue cohérente avec la frontière
    PALETTE = ["#00D4FF","#F0B429","#00C896","#FF4D6D","#9D7FEA",
               "#FF7849","#00B4D8","#F72585","#06D6A0","#FFD166",
               "#4CC9F0","#B5E48C","#06BEE1","#FFAA5B","#C77DFF"]

    with plt.rc_context(PLT_DARK):
        fig, axes = plt.subplots(1, 3, figsize=(13, 4.5))
        fig.patch.set_alpha(0)

        for ax, key, label, col in zip(axes, keys, labels_, accent):
            w    = opt_results[key]["weights"]
            mask = w > 0.005
            w_f  = w[mask]
            t_f  = [tickers[i] for i, m in enumerate(mask) if m]
            idx  = [i for i, m in enumerate(mask) if m]

            pie_colors = [PALETTE[i % len(PALETTE)] for i in idx]

            wedges, texts, autotexts = ax.pie(
                w_f,
                labels=t_f,
                autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
                colors=pie_colors,
                startangle=90,
                pctdistance=0.72,
                wedgeprops={"edgecolor": "#07090F", "linewidth": 1.8},
                textprops={"fontsize": 7.5},
            )
            for text in texts:
                text.set_color("#d0daea")
                text.set_fontsize(7.5)
            for at in autotexts:
                at.set_fontsize(7)
                at.set_color("#07090F")
                at.set_fontweight("bold")

            r, v, s = opt_results[key]["stats"]
            # Cercle central (donut effect)
            centre = plt.Circle((0, 0), 0.42, color="#07090F", zorder=10)
            ax.add_patch(centre)
            ax.text(0, 0.08, f"{s:.2f}", ha="center", va="center",
                    fontsize=14, fontweight="bold", color=col, zorder=11)
            ax.text(0, -0.14, "Sharpe", ha="center", va="center",
                    fontsize=7, color="#8aa0bc", zorder=11)

            # Halo coloré autour du cercle
            ring = plt.Circle((0, 0), 0.44, color=col, alpha=0.18, zorder=9)
            ax.add_patch(ring)

            ax.set_title(
                f"{label}\n"
                f"Rdt {r*100:.1f}%  ·  Vol {v*100:.1f}%",
                fontsize=9, color=col, pad=8, fontweight="bold"
            )
        plt.tight_layout()
        return fig


# ══════════════════════════════════════════════════════════════════════════════
# STRESS-TESTING
# ══════════════════════════════════════════════════════════════════════════════

def appliquer_stress(rendements_port: np.ndarray, pv: float,
                     scenario: dict, var_results: dict, alpha: float = 0.99) -> dict:
    mu   = rendements_port.mean()
    sig  = rendements_port.std()
    choc = scenario["choc_marche"] if scenario["choc_marche"] is not None else mu - 3*sig

    # P&L stressé
    pnl_stress = choc * pv

    # VaR stressée (vol multipliée)
    sigma_stress = sig * scenario["choc_vol"]
    z = stats.norm.ppf(1 - alpha)
    var_stress_vc = -z * sigma_stress * pv

    # Comparaison avec VaR normale
    var_normal = var_results.get("Variance-Covariance", {}).get(alpha, {}).get("VaR", 0)

    # Pire perte historique sur la période
    pire_perte = rendements_port.min() * pv

    return {
        "pnl_stress": pnl_stress,
        "var_stress":  var_stress_vc,
        "var_normal":  var_normal,
        "ratio":       var_stress_vc / var_normal if var_normal > 0 else np.nan,
        "pire_perte":  pire_perte,
        "choc":        choc,
        "choc_vol":    scenario["choc_vol"],
    }


def fig_stress_comparaison(stress_results: dict, pv: float) -> plt.Figure:
    with plt.rc_context(PLT_DARK):
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 4.5))
        fig.patch.set_alpha(0)

        scenarios = list(stress_results.keys())
        pnls    = [stress_results[s]["pnl_stress"]/1000 for s in scenarios]
        vars_nm = [stress_results[s]["var_normal"]/1000  for s in scenarios]
        ratios  = [stress_results[s]["ratio"] for s in scenarios]
        short_sc = [s.split("(")[0].strip()[:20] for s in scenarios]
        x = np.arange(len(scenarios))

        # ── Ax1 : barres horizontales ─────────────────────────────────────────
        bar_h1 = ax1.barh(x - 0.2, [abs(p) for p in pnls], 0.38,
                          color="#FF4D6D", alpha=0.85, label="P&L stressé")
        bar_h2 = ax1.barh(x + 0.2, vars_nm, 0.38,
                          color="#00D4FF", alpha=0.65, label="VaR normale 99%")

        # Labels valeurs
        for bar in bar_h1:
            w = bar.get_width()
            ax1.text(w + max(vars_nm)*0.02, bar.get_y() + bar.get_height()/2,
                     f"{w:.0f}k", va="center", fontsize=7.5,
                     color="#FF4D6D", fontweight="bold")

        ax1.set_yticks(x)
        ax1.set_yticklabels(short_sc, fontsize=8.5)
        ax1.set_xlabel("k€", fontsize=9, labelpad=6)
        ax1.set_title("Impact P&L vs VaR normale", fontsize=11,
                      color="#F0B429", pad=10, fontweight="bold")
        ax1.legend(fontsize=8.5, facecolor="#111827", edgecolor="#1a2235",
                   labelcolor="#d0daea")
        ax1.grid(axis="x", alpha=0.18, linestyle="--")
        for spine in ax1.spines.values(): spine.set_edgecolor("#1a2235")

        # ── Ax2 : multiplicateurs ─────────────────────────────────────────────
        colors_r = ["#e05252" if r > 2 else "#C9A84C" if r > 1.5 else "#1db87a"
                    for r in ratios]
        bars2 = ax2.bar(x, ratios, 0.6, color=colors_r, alpha=0.85,
                        edgecolor="#07090F", linewidth=0.8)

        # Labels sur les barres
        for bar, ratio in zip(bars2, ratios):
            ax2.text(bar.get_x() + bar.get_width()/2,
                     bar.get_height() + 0.04,
                     f"×{ratio:.2f}", ha="center", va="bottom",
                     fontsize=8, fontweight="bold",
                     color="#FF4D6D" if ratio > 2 else "#C9A84C" if ratio > 1.5 else "#1db87a")

        ax2.axhline(1.0, color="#00C896", lw=1.8, ls="--",
                    label="VaR normale (×1)", alpha=0.8)
        ax2.axhline(2.0, color="#FF4D6D", lw=1.4, ls=":",
                    label="Seuil alerte (×2)", alpha=0.8)
        # Zone d'alerte
        ax2.axhspan(2.0, max(ratios)*1.15 if max(ratios) > 2 else 2.5,
                    alpha=0.05, color="#FF4D6D")

        ax2.set_xticks(x)
        ax2.set_xticklabels(short_sc, rotation=30, ha="right", fontsize=8)
        ax2.set_ylabel("Ratio VaR Stressée / VaR Normale", fontsize=9, labelpad=6)
        ax2.set_title("Multiplicateurs de stress", fontsize=11,
                      color="#F0B429", pad=10, fontweight="bold")
        ax2.legend(fontsize=8.5, facecolor="#111827", edgecolor="#1a2235",
                   labelcolor="#d0daea")
        ax2.grid(axis="y", alpha=0.18, linestyle="--")
        for spine in ax2.spines.values(): spine.set_edgecolor("#1a2235")

        plt.tight_layout()
        return fig


# ══════════════════════════════════════════════════════════════════════════════
# MOTEUR VAR — 7 MÉTHODES (avec ES analytique TVE-GARCH)
# ══════════════════════════════════════════════════════════════════════════════

class VaREngine:
    def __init__(self, rendements: np.ndarray, pv: float = 10_000_000, horizon: int = 1):
        self.r       = rendements
        self.pv      = pv
        self.horizon = horizon
        self.n       = len(rendements)

    def historique(self, alpha: float) -> dict:
        q   = np.percentile(self.r, (1-alpha)*100)
        exc = self.r[self.r <= q]
        es  = exc.mean() if len(exc) > 0 else q
        return {"VaR": -q*self.pv*np.sqrt(self.horizon),
                "ES":  -es*self.pv*np.sqrt(self.horizon),
                "VaR_pct": -q}

    def variance_covariance(self, alpha: float) -> dict:
        mu, sigma = self.r.mean(), self.r.std()
        z   = stats.norm.ppf(1-alpha)
        var = -(mu + z*sigma) * np.sqrt(self.horizon)
        es  = (sigma*stats.norm.pdf(z)/(1-alpha) - mu) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"μ={mu*100:.4f}%, σ={sigma*100:.4f}%"}

    def riskmetrics(self, alpha: float, lam: float = 0.94) -> dict:
        r = self.r
        sig2 = np.zeros(len(r)); sig2[0] = r[0]**2
        for t in range(1, len(r)):
            sig2[t] = lam*sig2[t-1] + (1-lam)*r[t-1]**2
        sigma_t = np.sqrt(sig2[-1])
        z = stats.norm.ppf(1-alpha)
        var = -z * sigma_t * np.sqrt(self.horizon)
        es  = sigma_t * stats.norm.pdf(z) / (1-alpha) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"λ=0.94, σ_t={sigma_t*100:.4f}%"}

    def cornish_fisher(self, alpha: float) -> dict:
        mu, sigma = self.r.mean(), self.r.std()
        s = float(stats.skew(self.r))
        k = float(stats.kurtosis(self.r))
        z = stats.norm.ppf(1-alpha)
        z_cf = z + (z**2-1)*s/6 + (z**3-3*z)*k/24 - (2*z**3-5*z)*s**2/36
        var = -(mu + z_cf*sigma) * np.sqrt(self.horizon)
        es  = (sigma*(stats.norm.pdf(z_cf)/(1-alpha)) - mu) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"z_CF={z_cf:.4f}, skew={s:.3f}, kurt={k:.3f}"}

    def _fit_garch(self):
        r = self.r
        def neg_ll(p):
            w, a, b = p
            if w<=0 or a<=0 or b<=0 or a+b>=1: return 1e10
            sig2 = np.zeros(len(r)); sig2[0] = np.var(r)
            for t in range(1,len(r)):
                sig2[t] = w + a*r[t-1]**2 + b*sig2[t-1]
            return 0.5*np.sum(np.log(2*np.pi*sig2) + r**2/sig2)
        try:
            res = minimize(neg_ll, [1e-6,0.08,0.89], method='L-BFGS-B',
                           bounds=[(1e-8,None),(0.001,0.3),(0.5,0.999)])
            w, a, b = res.x
        except Exception:
            w, a, b = 5e-7, 0.09, 0.90
        sig2 = np.zeros(len(r)); sig2[0] = np.var(r)
        for t in range(1,len(r)):
            sig2[t] = w + a*r[t-1]**2 + b*sig2[t-1]
        return w, a, b, sig2

    def garch(self, alpha: float) -> dict:
        w, a, b, sig2 = self._fit_garch()
        r = self.r
        sigma_f = np.sqrt(w + a*r[-1]**2 + b*sig2[-1])
        z   = stats.norm.ppf(1-alpha)
        var = -z * sigma_f * np.sqrt(self.horizon)
        es  = sigma_f * stats.norm.pdf(z) / (1-alpha) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"ω={w:.2e}, α={a:.3f}, β={b:.3f}"}

    def _fit_gpd(self, excesses: np.ndarray):
        """Ajustement GPD par MLE."""
        def neg_ll_gpd(p):
            xi, beta = p
            if beta <= 0: return 1e10
            u_ = excesses / beta
            if xi != 0:
                if np.any(1+xi*u_ <= 0): return 1e10
                return len(u_)*np.log(beta) + (1+1/xi)*np.sum(np.log(1+xi*u_))
            return len(u_)*np.log(beta) + np.sum(u_)
        try:
            res = minimize(neg_ll_gpd, [0.1, np.std(excesses)], method='L-BFGS-B',
                           bounds=[(-0.5,0.5),(1e-6,None)])
            return res.x if res.success else np.array([0.1, np.std(excesses)])
        except Exception:
            return np.array([0.1, np.std(excesses)])

    def tve(self, alpha: float) -> dict:
        losses = -self.r
        u = np.percentile(losses, 90)
        exc = losses[losses > u] - u
        xi, beta = self._fit_gpd(exc)
        n_u, n = len(exc), len(losses)
        p = 1 - alpha
        if xi != 0:
            var = u + (beta/xi)*((n/n_u*p)**(-xi)-1)
        else:
            var = u + beta*np.log(n/n_u*p)
        var = max(var, 0)
        # ES analytique GPD
        if xi < 1:
            es = (var + beta - xi*u) / (1 - xi)
        else:
            es = var * 2
        return {"VaR": var*self.pv*np.sqrt(self.horizon),
                "ES":  es*self.pv*np.sqrt(self.horizon),
                "VaR_pct": var,
                "params": f"ξ={xi:.3f}, β={beta:.4f}, u={u:.4f}"}

    def tve_garch(self, alpha: float) -> dict:
        """TVE-GARCH avec ES analytique depuis la GPD."""
        r = self.r
        w, a, b, sig2 = self._fit_garch()
        # Résidus standardisés
        resid = r / np.sqrt(sig2)
        # Ajustement GPD sur les résidus
        losses_r = -resid
        u_r = np.percentile(losses_r, 90)
        exc_r = losses_r[losses_r > u_r] - u_r
        xi, beta_r = self._fit_gpd(exc_r)
        n_u, n = len(exc_r), len(losses_r)
        p = 1 - alpha
        if xi != 0:
            var_r = u_r + (beta_r/xi)*((n/n_u*p)**(-xi)-1)
        else:
            var_r = u_r + beta_r*np.log(n/n_u*p)
        var_r = max(var_r, 0)
        # ES analytique GPD
        if xi < 1:
            es_r = (var_r + beta_r - xi*u_r) / (1 - xi)
        else:
            es_r = var_r * 2
        # Volatilité prévisionnelle GARCH
        sigma_f = np.sqrt(w + a*r[-1]**2 + b*sig2[-1])
        var = var_r * sigma_f * np.sqrt(self.horizon)
        es  = es_r  * sigma_f * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"GARCH+GPD hybride, ξ={xi:.3f}, σ_f={sigma_f*100:.4f}%"}

    def compute_all(self, alphas=(0.95, 0.99)) -> dict:
        methods = {
            "Historique":          self.historique,
            "Variance-Covariance": self.variance_covariance,
            "RiskMetrics":         self.riskmetrics,
            "Cornish-Fisher":      self.cornish_fisher,
            "GARCH(1,1)":          self.garch,
            "TVE (POT)":           self.tve,
            "TVE-GARCH":           self.tve_garch,
        }
        results = {}
        for name, fn in methods.items():
            results[name] = {}
            for a in alphas:
                try:
                    results[name][a] = fn(a)
                except Exception as e:
                    results[name][a] = {"VaR": np.nan, "ES": np.nan,
                                        "VaR_pct": np.nan, "params": str(e)}
        return results


# ══════════════════════════════════════════════════════════════════════════════
# BACKTESTING
# ══════════════════════════════════════════════════════════════════════════════

def kupiec_test(rendements: np.ndarray, var_pct: float, alpha: float) -> dict:
    exc = (rendements < -var_pct).astype(int)
    N, T = int(exc.sum()), len(exc)
    if T == 0: return {"LR": np.nan, "p_value": np.nan, "valid": False, "N": 0, "T": 0}
    p0 = 1 - alpha
    p_hat = N / T
    if p_hat == 0:
        lr = -2 * T * np.log(1 - p0)
    elif p_hat == 1:
        lr = -2 * N * np.log(p0)
    else:
        lr = -2 * (T*np.log(1-p0) + N*np.log(p0)
                   - N*np.log(p_hat) - (T-N)*np.log(1-p_hat))
    pv = 1 - stats.chi2.cdf(max(lr,0), df=1)
    return {"LR": lr, "p_value": pv, "valid": pv > 0.05,
            "N": N, "T": T, "rate": p_hat, "expected": p0}


def christoffersen_test(rendements: np.ndarray, var_pct: float) -> dict:
    exc = (rendements < -var_pct).astype(int)
    n00 = np.sum((exc[:-1]==0) & (exc[1:]==0))
    n01 = np.sum((exc[:-1]==0) & (exc[1:]==1))
    n10 = np.sum((exc[:-1]==1) & (exc[1:]==0))
    n11 = np.sum((exc[:-1]==1) & (exc[1:]==1))
    pi01 = n01 / (n00+n01+1e-10)
    pi11 = n11 / (n10+n11+1e-10)
    pi   = (n01+n11) / (n00+n01+n10+n11+1e-10)
    try:
        lr = -2*(
            (n00+n10)*np.log(max(1-pi,1e-15))+(n01+n11)*np.log(max(pi,1e-15))
            - n00*np.log(max(1-pi01,1e-15)) - n01*np.log(max(pi01,1e-15))
            - n10*np.log(max(1-pi11,1e-15)) - n11*np.log(max(pi11,1e-15))
        )
    except Exception:
        lr = np.nan
    pv = 1 - stats.chi2.cdf(max(lr,0), df=1) if not np.isnan(lr) else np.nan
    return {"LR_ind": lr, "p_value_ind": pv,
            "valid": pv > 0.05 if not np.isnan(pv) else False}


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS GRAPHIQUES
# ══════════════════════════════════════════════════════════════════════════════

PLT_DARK = {
    "figure.facecolor": "#07090F", "axes.facecolor": "#0C0F1A",
    "axes.edgecolor": "#1a2235",   "axes.labelcolor": "#7A8BA8",
    "xtick.color": "#7A8BA8",      "ytick.color": "#7A8BA8",
    "grid.color":  "#111827",      "text.color":  "#E8EDF5",
    "legend.facecolor": "#111827", "legend.edgecolor": "#1a2235",
}

def fig_perf(rendements: pd.DataFrame, tickers: list) -> plt.Figure:
    with plt.rc_context(PLT_DARK):
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 5.5),
                                        gridspec_kw={"height_ratios": [3, 1]})
        fig.patch.set_alpha(0)
        colors = ["#00D4FF","#F0B429","#00C896","#FF4D6D",
                  "#9D7FEA","#FF7849","#4CC9F0","#F72585"]
        port_r = rendements.mean(axis=1)
        cumul  = (1 + port_r).cumprod()

        ax1.fill_between(range(len(cumul)), cumul.values, 1,
                         where=(cumul.values >= 1),
                         alpha=0.12, color="#00D4FF", interpolate=True)
        ax1.fill_between(range(len(cumul)), cumul.values, 1,
                         where=(cumul.values < 1),
                         alpha=0.12, color="#FF4D6D", interpolate=True)
        ax1.plot(cumul.values, color="#00D4FF", lw=2, label="Portefeuille", zorder=4)
        for i, t in enumerate(tickers[:5]):
            if t in rendements.columns:
                c = (1 + rendements[t]).cumprod()
                ax1.plot(c.values, color=colors[(i+1) % len(colors)],
                         lw=0.9, alpha=0.55, label=t, zorder=3)
        ax1.axhline(1, color="#F0B429", lw=0.8, ls="--", alpha=0.5)
        ax1.set_title("Performance cumulée", fontsize=11, color="#F0B429",
                      pad=8, fontweight="bold")
        ax1.legend(fontsize=7.5, loc="upper left",
                   facecolor="#111827", edgecolor="#1a2235", labelcolor="#d0daea")
        ax1.grid(True, alpha=0.15, linestyle="--")
        ax1.tick_params(labelsize=8.5)
        for spine in ax1.spines.values(): spine.set_edgecolor("#1a2235")

        col_bars = ["#1db87a" if v >= 0 else "#e05252" for v in port_r]
        ax2.bar(range(len(port_r)), port_r * 100, color=col_bars,
                alpha=0.75, width=1, linewidth=0)
        ax2.axhline(0, color="#F0B429", lw=0.6)
        ax2.set_title("Rendements journaliers (%)", fontsize=9,
                      color="#8aa0bc", pad=4)
        ax2.tick_params(labelsize=7.5)
        ax2.grid(True, alpha=0.15, linestyle="--")
        for spine in ax2.spines.values(): spine.set_edgecolor("#1a2235")
        plt.tight_layout()
        return fig


def fig_correlation(rendements: pd.DataFrame) -> plt.Figure:
    corr = rendements.corr()
    tickers = list(corr.columns)
    n = len(tickers)
    sz = max(6, n * 1.05 + 1.5)
    with plt.rc_context(PLT_DARK):
        fig, ax = plt.subplots(figsize=(sz, sz * 0.82))
        fig.patch.set_alpha(0)
        from matplotlib.colors import LinearSegmentedColormap
        cmap_custom = LinearSegmentedColormap.from_list(
            "var_corr", ["#e05252", "#111827", "#1db87a"], N=256)
        im = ax.imshow(corr.values, cmap=cmap_custom, vmin=-1, vmax=1, aspect="auto")
        cbar = plt.colorbar(im, ax=ax, shrink=0.82, pad=0.02)
        cbar.set_label("Corrélation", fontsize=9, color="#8aa0bc")
        cbar.ax.yaxis.set_tick_params(color="#8aa0bc", labelsize=8)
        plt.setp(cbar.ax.yaxis.get_ticklabels(), color="#8aa0bc")
        ax.set_xticks(range(n)); ax.set_yticks(range(n))
        ax.set_xticklabels(tickers, rotation=40, ha="right", fontsize=8.5)
        ax.set_yticklabels(tickers, fontsize=8.5)
        for i in range(n):
            for j in range(n):
                v = corr.values[i, j]
                fc = "white" if abs(v) > 0.45 else "#d0daea"
                ax.text(j, i, f"{v:.2f}", ha="center", va="center",
                        fontsize=7.5, color=fc, fontweight="bold")
        ax.set_title("Matrice de Corrélation", fontsize=12,
                     color="#F0B429", pad=10, fontweight="bold")
        for spine in ax.spines.values(): spine.set_edgecolor("#1a2235")
        plt.tight_layout()
        return fig


def fig_var_comparaison(var_results: dict, conf: float, pv: float) -> plt.Figure:
    methods = list(var_results.keys())
    vars_   = [var_results[m][conf]["VaR"] / 1000 for m in methods]
    ess_    = [var_results[m][conf]["ES"]  / 1000 for m in methods]
    colors  = ["#00D4FF","#F0B429","#00C896","#FF4D6D",
               "#9D7FEA","#FF7849","#4CC9F0"]
    short   = [m.replace("Variance-Covariance", "VCV")
                .replace("Cornish-Fisher", "C-Fisher")
                .replace("RiskMetrics", "RiskM.") for m in methods]
    with plt.rc_context(PLT_DARK):
        fig, ax = plt.subplots(figsize=(12, 4.5))
        fig.patch.set_alpha(0)
        x = np.arange(len(methods)); w = 0.38
        bars1 = ax.bar(x - w/2, vars_, w, color=colors,
                       alpha=0.9, label="VaR", edgecolor="#07090F", linewidth=0.6)
        ax.bar(x + w/2, ess_, w, color=colors,
               alpha=0.42, label="ES (CVaR)",
               edgecolor="#07090F", linewidth=0.6, hatch="//")
        for bar, col in zip(bars1, colors):
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h + max(vars_)*0.015,
                    f"{h:.0f}k", ha="center", va="bottom",
                    fontsize=7.5, color=col, fontweight="bold")
        ax.set_xticks(x)
        ax.set_xticklabels(short, rotation=22, ha="right", fontsize=9)
        ax.set_ylabel("k€", fontsize=10, labelpad=6)
        ax.grid(axis="y", alpha=0.18, linestyle="--")
        ax.set_title(
            f"VaR & Expected Shortfall — Niveau {conf*100:.0f}%  ·  "
            f"Portefeuille {pv/1e6:.0f} M€",
            fontsize=11, color="#F0B429", pad=10, fontweight="bold")
        ax.legend(fontsize=9, facecolor="#111827", edgecolor="#1a2235",
                  labelcolor="#d0daea")
        for spine in ax.spines.values(): spine.set_edgecolor("#1a2235")
        ax.tick_params(labelsize=8.5)
        plt.tight_layout()
        return fig


def fig_distribution(rendements: np.ndarray, var_results: dict) -> plt.Figure:
    with plt.rc_context(PLT_DARK):
        fig, ax = plt.subplots(figsize=(12, 4.5))
        fig.patch.set_alpha(0)
        pos_mask = rendements >= 0
        ax.hist(rendements[pos_mask] * 100, bins=55, density=True,
                color="#00C896", alpha=0.45, edgecolor="#07090F",
                lw=0.2, label="Rdt ≥ 0")
        ax.hist(rendements[~pos_mask] * 100, bins=55, density=True,
                color="#00D4FF", alpha=0.55, edgecolor="#07090F",
                lw=0.2, label="Rdt < 0")
        mu, sig = rendements.mean(), rendements.std()
        x_range = np.linspace(rendements.min(), rendements.max(), 400)
        ax.plot(x_range * 100, stats.norm.pdf(x_range, mu, sig) / 100,
                color="#F0B429", lw=2.2, ls="--", label="N(μ, σ)", zorder=5)
        var_styles = {
            "Historique":  ("#e05252", "--"),
            "GARCH(1,1)": ("#a855f7", "-."),
            "TVE-GARCH":  ("#f97316", ":"),
        }
        for meth, (col, ls) in var_styles.items():
            if meth in var_results:
                p = var_results[meth].get(0.99, {}).get("VaR_pct")
                if p and not np.isnan(p):
                    ax.axvline(-p * 100, color=col, lw=1.8, ls=ls,
                               label=f"VaR 99% {meth}", zorder=6)
        ax.set_xlabel("Rendement journalier (%)", fontsize=10, labelpad=6)
        ax.set_ylabel("Densité", fontsize=10, labelpad=6)
        ax.set_title("Distribution des rendements  ·  VaR 99% superposée",
                     fontsize=11, color="#F0B429", pad=10, fontweight="bold")
        ax.legend(fontsize=8, facecolor="#111827", edgecolor="#1a2235",
                  labelcolor="#d0daea")
        ax.grid(True, alpha=0.15, linestyle="--")
        for spine in ax.spines.values(): spine.set_edgecolor("#1a2235")
        ax.tick_params(labelsize=8.5)
        plt.tight_layout()
        return fig


def fig_backtesting(bt_results: dict) -> plt.Figure:
    methods = list(bt_results.keys())
    short   = [m.replace("Variance-Covariance", "VCV")
                .replace("Cornish-Fisher", "C-Fisher")
                .replace("RiskMetrics", "RiskM.") for m in methods]
    p95 = [bt_results[m][0.95]["p_value"] for m in methods if 0.95 in bt_results[m]]
    p99 = [bt_results[m][0.99]["p_value"] for m in methods if 0.99 in bt_results[m]]
    with plt.rc_context(PLT_DARK):
        fig, axes = plt.subplots(1, 2, figsize=(12, 4))
        fig.patch.set_alpha(0)
        for ax, pvals, title in zip(axes,
                                    [p95, p99],
                                    ["Test de Kupiec — 95%", "Test de Kupiec — 99%"]):
            x = np.arange(len(short[:len(pvals)]))
            for xi, (pv_, col) in enumerate(zip(pvals,
                    ["#1db87a" if p > 0.05 else "#e05252" for p in pvals])):
                ax.bar(xi, pv_, 0.6, color=col, alpha=0.85,
                       edgecolor="#07090F", linewidth=0.7)
                ax.text(xi, pv_ + 0.008, f"{pv_:.3f}",
                        ha="center", va="bottom", fontsize=7.5,
                        color=col, fontweight="bold")
            ax.axhline(0.05, color="#F0B429", lw=2, ls="--",
                       label="Seuil α = 5%", alpha=0.9)
            ax.axhspan(0, 0.05, alpha=0.06, color="#FF4D6D")
            ax.set_xticks(x)
            ax.set_xticklabels(short[:len(pvals)],
                               rotation=28, ha="right", fontsize=8.5)
            ax.set_ylabel("p-value", fontsize=10, labelpad=6)
            ax.set_ylim(0, max(max(pvals) * 1.2, 0.15))
            ax.set_title(title, fontsize=11, color="#F0B429",
                         pad=8, fontweight="bold")
            ax.legend(fontsize=9, facecolor="#111827", edgecolor="#1a2235",
                      labelcolor="#d0daea")
            ax.grid(axis="y", alpha=0.15, linestyle="--")
            for spine in ax.spines.values(): spine.set_edgecolor("#1a2235")
            ax.tick_params(labelsize=8.5)
        plt.tight_layout()
        return fig


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generer_excel(rendements_df: pd.DataFrame, var_results: dict,
                   bt_results: dict, pv: float, opt_results: dict = None) -> bytes | None:
    if not HAS_XLSX: return None
    wb = Workbook()
    NAVY, BLUE, GOLD = "1B2A4A", "2E5FA3", "C9A84C"
    WHITE, LGRAY = "FFFFFF", "F0F4FA"

    def th(ws, r, c, v, bg=NAVY, fg=WHITE):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="Calibri", bold=True, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        s = Side(border_style="thin", color="CCCCCC")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    def td(ws, r, c, v, bg=WHITE):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="Calibri", size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        s = Side(border_style="thin", color="DDDDDD")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    # Feuille 1 : Résumé VaR
    ws1 = wb.active; ws1.title = "📊 Résumé VaR"
    ws1.merge_cells("A1:F1")
    c = ws1["A1"]; c.value = "RAPPORT VaR — SYNTHÈSE DES RÉSULTATS"
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30
    headers = ["Méthode","VaR 95% (€)","VaR 95% (%)","VaR 99% (€)","VaR 99% (%)","ES 99% (€)"]
    for j,h in enumerate(headers,1): th(ws1, 2, j, h, bg=BLUE)
    for i,(m,res) in enumerate(var_results.items(), 3):
        r95 = res.get(0.95,{}); r99 = res.get(0.99,{})
        bg = LGRAY if i%2==0 else WHITE
        td(ws1,i,1,m,bg=bg)
        td(ws1,i,2,round(r95.get("VaR",0),0),bg=bg)
        td(ws1,i,3,f"{r95.get('VaR_pct',0)*100:.3f}%",bg=bg)
        td(ws1,i,4,round(r99.get("VaR",0),0),bg=bg)
        td(ws1,i,5,f"{r99.get('VaR_pct',0)*100:.3f}%",bg=bg)
        td(ws1,i,6,round(r99.get("ES",0),0),bg=bg)
    for w,col in zip([26,16,12,16,12,16],["A","B","C","D","E","F"]):
        ws1.column_dimensions[col].width = w

    # Feuille 2 : Backtesting
    ws2 = wb.create_sheet("🧪 Backtesting")
    hdrs = ["Méthode","CL","Exceptions","T","Taux obs.","Taux att.",
            "Kupiec LR","Kupiec p","Kupiec OK","CC LR","CC p","CC OK"]
    for j,h in enumerate(hdrs,1): th(ws2,1,j,h,bg=BLUE)
    row = 2
    for m, alphas in bt_results.items():
        for a, res in alphas.items():
            bg = LGRAY if row%2==0 else WHITE
            k = res.get("kupiec",{}); cc = res.get("cc",{})
            td(ws2,row,1,m,bg=bg); td(ws2,row,2,f"{a*100:.0f}%",bg=bg)
            td(ws2,row,3,k.get("N",""),bg=bg); td(ws2,row,4,k.get("T",""),bg=bg)
            td(ws2,row,5,f"{k.get('rate',0)*100:.2f}%",bg=bg)
            td(ws2,row,6,f"{(1-a)*100:.2f}%",bg=bg)
            td(ws2,row,7,round(k.get("LR",0),3),bg=bg)
            td(ws2,row,8,round(k.get("p_value",0),4),bg=bg)
            td(ws2,row,9,"OUI" if k.get("valid") else "NON",bg=bg)
            td(ws2,row,10,round(cc.get("LR_ind",0),3),bg=bg)
            td(ws2,row,11,round(cc.get("p_value_ind",0),4),bg=bg)
            td(ws2,row,12,"OUI" if cc.get("valid") else "NON",bg=bg)
            row += 1
    for w,col in zip([26,6,12,8,10,10,10,10,10,10,10,10],
                     ["A","B","C","D","E","F","G","H","I","J","K","L"]):
        ws2.column_dimensions[col].width = w

    # Feuille 3 : Markowitz (si disponible)
    if opt_results:
        ws3 = wb.create_sheet("📐 Markowitz")
        th(ws3,1,1,"Actif",bg=BLUE); th(ws3,1,2,"Sharpe Max. (%)",bg=BLUE)
        th(ws3,1,3,"Variance Min. (%)",bg=BLUE); th(ws3,1,4,"Équipondéré (%)",bg=BLUE)
        for i, ticker in enumerate(opt_results["tickers"], 2):
            bg = LGRAY if i%2==0 else WHITE
            td(ws3,i,1,ticker,bg=bg)
            td(ws3,i,2,f"{opt_results['sharpe']['weights'][i-2]*100:.1f}%",bg=bg)
            td(ws3,i,3,f"{opt_results['minvar']['weights'][i-2]*100:.1f}%",bg=bg)
            td(ws3,i,4,f"{opt_results['equi']['weights'][i-2]*100:.1f}%",bg=bg)
        row_stats = len(opt_results["tickers"]) + 3
        th(ws3,row_stats,1,"Statistiques",bg=NAVY,fg=WHITE)
        th(ws3,row_stats,2,"Sharpe Max.",bg=NAVY,fg=WHITE)
        th(ws3,row_stats,3,"Variance Min.",bg=NAVY,fg=WHITE)
        th(ws3,row_stats,4,"Équipondéré",bg=NAVY,fg=WHITE)
        for j,key in enumerate(["sharpe","minvar","equi"],2):
            r,v,s = opt_results[key]["stats"]
            td(ws3,row_stats+1,j,f"{r*100:.2f}%")
            td(ws3,row_stats+2,j,f"{v*100:.2f}%")
            td(ws3,row_stats+3,j,f"{s:.3f}")
        for r_,label in zip([row_stats+1,row_stats+2,row_stats+3],
                             ["Rdt Annualisé","Vol. Annualisée","Ratio Sharpe"]):
            td(ws3,r_,1,label)
        ws3.column_dimensions["A"].width = 18
        for col in ["B","C","D"]: ws3.column_dimensions[col].width = 18

    # Feuille 4 : Rendements
    ws4 = wb.create_sheet("📋 Données")
    r_port = rendements_df.mean(axis=1).tail(250)
    for j,h in enumerate(["Date","Rdt Portfolio (%)"],1): th(ws4,1,j,h,bg=BLUE)
    for i,(d,v) in enumerate(r_port.items(),2):
        bg = LGRAY if i%2==0 else WHITE
        td(ws4,i,1,d.strftime("%d/%m/%Y"),bg=bg)
        td(ws4,i,2,round(v*100,4),bg=bg)
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 18

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT PDF
# ══════════════════════════════════════════════════════════════════════════════

def generer_pdf(var_results: dict, bt_results: dict, metrics: dict, pv: float,
                fig_var=None, fig_dist=None) -> bytes | None:
    if not HAS_PDF: return None
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    NAVY_C  = HexColor("#1B2A4A"); BLUE_C = HexColor("#2E5FA3")
    GOLD_C  = HexColor("#C9A84C"); LGRAY_C= HexColor("#F0F4FA")

    S = lambda name, **kw: ParagraphStyle(name, **kw)
    s_title = S("t", fontName="Helvetica-Bold", fontSize=22, textColor=HexColor("#FFFFFF"),
                alignment=TA_CENTER, spaceAfter=4)
    s_sub   = S("s", fontName="Helvetica-Oblique", fontSize=11, textColor=GOLD_C,
                alignment=TA_CENTER, spaceAfter=6)
    s_h2    = S("h2", fontName="Helvetica-Bold", fontSize=12, textColor=NAVY_C,
                spaceBefore=12, spaceAfter=6)
    s_body  = S("b", fontName="Helvetica", fontSize=9, textColor=HexColor("#3A3A3A"),
                alignment=TA_JUSTIFY, spaceAfter=6, leading=14)

    def tbl_style():
        return TableStyle([
            ("BACKGROUND",(0,0),(-1,0),BLUE_C),("TEXTCOLOR",(0,0),(-1,0),HexColor("#FFFFFF")),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8.5),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[HexColor("#FFFFFF"),LGRAY_C]),
            ("GRID",(0,0),(-1,-1),0.3,HexColor("#CCCCCC")),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ])

    story = []
    story.append(Spacer(1,2*cm))
    cover = Table([[Paragraph("RAPPORT DE GESTION DES RISQUES", s_title)]], [15*cm])
    cover.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),NAVY_C),
                                ("TOPPADDING",(0,0),(-1,-1),18),("BOTTOMPADDING",(0,0),(-1,-1),18)]))
    story.append(cover)
    story.append(Spacer(1,0.3*cm))
    cover2 = Table([[Paragraph("Value at Risk — 7 méthodes · Backtesting · Stress-Testing · Markowitz", s_sub)]],[15*cm])
    cover2.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),HexColor("#243B60")),
                                  ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    story.append(cover2)
    story.append(Spacer(1,1.5*cm))

    from datetime import date
    info = [["Valeur portefeuille", f"{pv:,.0f} €"],
            ["Horizon", "1 jour ouvré"],
            ["Niveaux de confiance", "95% et 99%"],
            ["Date de production", date.today().strftime("%d/%m/%Y")]]
    t_info = Table([[Paragraph(k,s_body),Paragraph(v,s_body)] for k,v in info], [5*cm,10*cm])
    t_info.setStyle(TableStyle([("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
                                  ("ROWBACKGROUNDS",(0,0),(-1,-1),[HexColor("#FFFFFF"),LGRAY_C]),
                                  ("GRID",(0,0),(-1,-1),0.3,HexColor("#DDDDDD")),
                                  ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.append(t_info); story.append(PageBreak())

    story.append(Paragraph("1. RÉSULTATS DE LA VALUE AT RISK", s_h2))
    story.append(HRFlowable(width="100%",thickness=1,color=GOLD_C,spaceAfter=8))
    var_data = [["Méthode","VaR 95% (€)","VaR 95% (%)","VaR 99% (€)","VaR 99% (%)","ES 99% (€)"]]
    for m, res in var_results.items():
        r95, r99 = res.get(0.95,{}), res.get(0.99,{})
        var_data.append([m,
            f"{r95.get('VaR',0):,.0f} €", f"{r95.get('VaR_pct',0)*100:.3f}%",
            f"{r99.get('VaR',0):,.0f} €", f"{r99.get('VaR_pct',0)*100:.3f}%",
            f"{r99.get('ES',0):,.0f} €"])
    t_var = Table(var_data, [3.5*cm,2.5*cm,2*cm,2.5*cm,2*cm,2.5*cm])
    t_var.setStyle(tbl_style()); story.append(t_var); story.append(Spacer(1,0.5*cm))

    if fig_var:
        buf_img = io.BytesIO(); fig_var.savefig(buf_img, format="png", dpi=120, bbox_inches="tight"); buf_img.seek(0)
        story.append(Image(buf_img, width=15*cm, height=5*cm))
        plt.close(fig_var)
    story.append(PageBreak())

    story.append(Paragraph("2. BACKTESTING", s_h2))
    story.append(HRFlowable(width="100%",thickness=1,color=GOLD_C,spaceAfter=8))
    story.append(Paragraph(
        "Test de Kupiec (POF) : H₀ → fréquence observée = 1−α. "
        "Test de Christoffersen : teste l'indépendance temporelle des exceptions. "
        "p > 0.05 → modèle non rejeté.", s_body))
    bt_data = [["Méthode","CL","Exceptions","Taux obs.","Kupiec p","Kupiec","CC p","CC"]]
    for m, alphas in bt_results.items():
        for a, res in alphas.items():
            k = res.get("kupiec",{}); cc = res.get("cc",{})
            bt_data.append([m, f"{a*100:.0f}%", str(k.get("N","")),
                f"{k.get('rate',0)*100:.2f}%",
                f"{k.get('p_value',0):.4f}", "OUI" if k.get("valid") else "NON",
                f"{cc.get('p_value_ind',0):.4f}", "OUI" if cc.get("valid") else "NON"])
    t_bt = Table(bt_data, [3.5*cm,1.2*cm,1.5*cm,1.8*cm,1.8*cm,1.5*cm,1.8*cm,1.5*cm])
    t_bt.setStyle(tbl_style()); story.append(t_bt); story.append(Spacer(1,0.5*cm))

    if fig_dist:
        buf_img2 = io.BytesIO(); fig_dist.savefig(buf_img2, format="png", dpi=120, bbox_inches="tight"); buf_img2.seek(0)
        story.append(Image(buf_img2, width=15*cm, height=5*cm))
        plt.close(fig_dist)

    doc.build(story)
    buf.seek(0); return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════

for key in ["prix","rendements","var_results","bt_results","actifs_choisis","pv","opt_results","stress_results"]:
    if key not in st.session_state:
        st.session_state[key] = None

# FIX 3 — Défauts stables pour le multiselect actifs (évite le bug "0 actif" au 1er rendu)
_ACTIFS_DEFAULT = ["Apple (AAPL)", "Microsoft (MSFT)", "LVMH (MC.PA)",
                   "TotalEnergies (TTE)", "BNP Paribas (BNP)"]
if "actifs_selection" not in st.session_state:
    st.session_state["actifs_selection"] = _ACTIFS_DEFAULT


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 📉 VaR Analytics Suite")
    st.markdown("<div style='font-size:10px;color:var(--txt-secondary,#7A8BA8);font-family:DM Mono,monospace;margin-bottom:16px'>v4.0 · Département Risque</div>", unsafe_allow_html=True)
    st.divider()

    menu = st.selectbox(
        "Navigation",
        ["🏠 Accueil", "🏦 Portefeuille", "📐 Optimisation", "📉 Calcul VaR",
         "🧪 Backtesting", "🔥 Stress-Testing", "📊 Reporting"],
        label_visibility="collapsed"
    )
    st.divider()

    if st.session_state["rendements"] is not None:
        r = st.session_state["rendements"].mean(axis=1)
        ann_r = r.mean() * 252
        ann_v = r.std() * np.sqrt(252)
        sharpe = (r.mean() - 0.03/252) / r.std() * np.sqrt(252)
        st.markdown("**Portefeuille chargé**")
        st.markdown(f"""
        <div style='font-size:11px;font-family:DM Mono,monospace;line-height:1.9'>
        <span style='color:var(--txt-secondary,#7A8BA8)'>Rdt ann. :</span> <span style='color:var(--signal-green,#00C896)'>+{ann_r*100:.2f}%</span><br>
        <span style='color:var(--txt-secondary,#7A8BA8)'>Vol. ann. :</span> <span style='color:var(--txt-primary,#E8EDF5)'>{ann_v*100:.2f}%</span><br>
        <span style='color:var(--txt-secondary,#7A8BA8)'>Sharpe   :</span> <span style='color:var(--amber,#F0B429)'>{sharpe:.3f}</span>
        </div>""", unsafe_allow_html=True)
        st.divider()

    st.markdown("""
    <div style='font-size:10px;color:var(--txt-secondary,#7A8BA8);line-height:1.8'>
    <b style='color:var(--amber,#F0B429)'>7 méthodes VaR</b><br>
    · Historique · VCV · RiskMetrics<br>
    · Cornish-Fisher · GARCH(1,1)<br>
    · TVE (POT) · TVE-GARCH<br><br>
    <b style='color:var(--amber,#F0B429)'>Nouveautés v4.0</b><br>
    · Optimisation Markowitz<br>
    · Stress-Testing (6 scénarios)<br>
    · ES analytique TVE-GARCH<br>
    · Matrice corrélation
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : ACCUEIL
# ══════════════════════════════════════════════════════════════════════════════

if menu == "🏠 Accueil":
    st.title("VaR Analytics Suite")
    st.markdown("""
    <div class='info-box'>
    Progiciel professionnel de calcul, comparaison et validation de la <b>Value at Risk</b>
    sur un portefeuille d'actions. Développé selon les standards <b>Bâle III/IV</b>.
    Version 4.0 : optimisation Markowitz, stress-testing et ES analytique intégrés.
    </div>""", unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("Méthodes VaR", "7", "Complètes")
    with col2: st.metric("Tests backtest", "2", "Kupiec + CC")
    with col3: st.metric("Scénarios Stress", "6", "Historiques")
    with col4: st.metric("Export", "Excel + PDF", "4 feuilles")

    st.markdown("<div class='section-header'>Fonctionnalités</div>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
        **📥 Données de marché**
        - Téléchargement automatique Yahoo Finance
        - 15 actifs prédéfinis + simulation intégrée
        - Matrice de corrélation interactive

        **📐 Optimisation Markowitz** *(nouveau)*
        - Frontière efficiente complète
        - 3 portefeuilles : Sharpe Max, Variance Min, Équipondéré
        - Allocation recommandée exportable
        """)
    with c2:
        st.markdown("""
        **🔥 Stress-Testing** *(nouveau)*
        - 6 scénarios historiques (2008, COVID, Brexit…)
        - Comparaison VaR normale vs stressée
        - Multiplicateurs et seuils d'alerte

        **📊 Reporting amélioré**
        - Excel 4 feuilles (+ feuille Markowitz)
        - PDF exécutif avec graphiques
        - Prêt Direction Risque Bâle III/IV
        """)

    st.markdown("<div class='section-header'>Équipe Projet</div>", unsafe_allow_html=True)
    cols = st.columns(4)
    membres = ["Anta Mbaye", "Harlem D. Adjagba", "Ecclésiaste Gnargo", "Wariol G. Kopangoye"]
    for col, m in zip(cols, membres):
        with col:
            st.markdown(f"""
            <div class='var-card' style='text-align:center;padding:14px'>
            <div style='font-size:22px;margin-bottom:6px'>👤</div>
            <div style='font-size:11px;font-weight:600;color:var(--txt-primary,#E8EDF5)'>{m}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style='text-align:center;margin-top:16px;font-size:11px;color:var(--txt-secondary,#7A8BA8)'>
    Double diplôme M2 IFIM · Ing 3 MACS — Mathématiques Appliquées au Calcul Scientifique
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : PORTEFEUILLE
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "🏦 Portefeuille":
    st.title("Construction du Portefeuille")

    st.markdown("<div class='section-header'>Sélection des actifs</div>", unsafe_allow_html=True)
    col1, col2 = st.columns([2,1])
    with col1:
        actifs_choisis = st.multiselect(
            "Actifs financiers",
            list(ACTIFS_DISPONIBLES.keys()),
            default=st.session_state["actifs_selection"],
            key="actifs_selection",
        )
    with col2:
        pv_millions = st.number_input("Valeur du portefeuille (M€)", min_value=0.1, max_value=1000.0, value=10.0, step=0.5)
        pv = pv_millions * 1_000_000

    col3, col4, col5 = st.columns(3)
    with col3: date_debut = st.date_input("Date de début", value=pd.to_datetime("2019-01-01"))
    with col4: date_fin   = st.date_input("Date de fin",   value=pd.to_datetime("today"))
    with col5:
        source = st.radio("Source données", ["Yahoo Finance", "Simulation"], horizontal=True)

    btn_col, _ = st.columns([1,3])
    with btn_col:
        btn = st.button("▶  Charger les données", type="primary", use_container_width=True)

    # Sync pv dans session_state dès que l'input change (sans attendre le bouton)
    if st.session_state.get("pv") != pv:
        st.session_state["pv"] = pv

    if btn:
        if len(actifs_choisis) < 2:
            st.warning("Sélectionnez au moins 2 actifs.")
        elif date_debut >= date_fin:
            st.warning("La date de début doit être antérieure à la date de fin.")
        else:
            tickers = [ACTIFS_DISPONIBLES[a] for a in actifs_choisis]
            with st.spinner("Chargement des données en cours…"):
                if source == "Yahoo Finance" and HAS_YF:
                    # Passer les dates en string pour éviter les problèmes de cache
                    prix = telecharger_donnees(tickers, str(date_debut), str(date_fin))
                    if prix.empty:
                        st.warning("Aucune donnée récupérée, passage en simulation.")
                        prix = donnees_simulation(tickers)
                else:
                    prix = donnees_simulation(tickers)

                # Calcul des rendements log (plus stable que pct_change pour de longues séries)
                prix_clean = prix.ffill().bfill()
                rendements = np.log(prix_clean / prix_clean.shift(1)).iloc[1:]

                # Nettoyage robuste : supprimer les colonnes avec > 5% de NaN seulement
                seuil_nan = 0.05
                cols_ok = rendements.columns[rendements.isna().mean() < seuil_nan]
                if len(cols_ok) < 2:
                    # Si trop de colonnes supprimées, utiliser un seuil plus souple
                    cols_ok = rendements.columns[rendements.isna().mean() < 0.30]
                rendements = rendements[cols_ok].dropna()
                prix = prix_clean[cols_ok]
                tickers_valides = list(rendements.columns)

                # Vérification finale
                if rendements.empty or len(tickers_valides) < 2:
                    st.error("Données insuffisantes après nettoyage. Essayez une période plus longue ou passez en mode Simulation.")
                    st.stop()

            st.session_state.update({
                "prix": prix, "rendements": rendements,
                "actifs_choisis": tickers_valides, "pv": pv,
                "var_results": None, "bt_results": None,
                "opt_results": None, "stress_results": None
            })
            n_jours = len(rendements)
            n_actifs = len(tickers_valides)
            supprimes = len(tickers) - n_actifs
            msg = f"✅ {n_actifs} actif(s) chargés — {n_jours} jours de données."
            if supprimes > 0:
                msg += f" ({supprimes} actif(s) retirés pour données insuffisantes)"
            st.success(msg)

    if st.session_state["rendements"] is not None:
        rendements = st.session_state["rendements"]
        prix       = st.session_state["prix"]

        st.markdown("<div class='section-header'>Statistiques du portefeuille</div>", unsafe_allow_html=True)
        port_r = rendements.mean(axis=1).dropna()
        ann_r  = port_r.mean() * 252
        ann_v  = port_r.std()  * np.sqrt(252)
        sharpe = ((port_r.mean() - 0.03/252) / port_r.std() * np.sqrt(252)) if port_r.std() > 0 else 0.0
        skew   = float(stats.skew(port_r)) if len(port_r) > 3 else 0.0
        kurt   = float(stats.kurtosis(port_r)) if len(port_r) > 3 else 0.0
        cum    = (1 + port_r).cumprod()
        mdd    = float((cum / cum.cummax() - 1).min()) if len(cum) > 0 else 0.0

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Rdt Annualisé",  f"+{ann_r*100:.2f}%")
        c2.metric("Vol. Annuelle",   f"{ann_v*100:.2f}%")
        c3.metric("Sharpe",          f"{sharpe:.3f}")
        c4.metric("Max Drawdown",    f"{mdd*100:.2f}%")
        c5.metric("Skewness",        f"{skew:.4f}")
        c6.metric("Kurtosis (exc)",  f"{kurt:.4f}")

        st.markdown("<div class='section-header'>Performance & Rendements</div>", unsafe_allow_html=True)
        fig = fig_perf(rendements, list(rendements.columns))
        st.pyplot(fig, use_container_width=True); plt.close(fig)

        st.markdown("<div class='section-header'>Matrice de Corrélation</div>", unsafe_allow_html=True)
        fig_corr = fig_correlation(rendements)
        st.pyplot(fig_corr, use_container_width=True); plt.close(fig_corr)

        st.markdown("<div class='section-header'>Statistiques individuelles</div>", unsafe_allow_html=True)
        def _safe_stat(fn, col):
            try:
                v = fn(col.dropna())
                return round(float(v), 4) if np.isfinite(v) else 0.0
            except Exception:
                return 0.0

        stats_df = pd.DataFrame({
            "Secteur":      [SECTEURS.get(t, "—") for t in rendements.columns],
            "Rdt moy. (%)": (rendements.mean() * 252 * 100).round(2).values,
            "Vol. ann. (%)": (rendements.std() * np.sqrt(252) * 100).round(2).values,
            "Skewness":     [_safe_stat(stats.skew, rendements[c]) for c in rendements.columns],
            "Kurtosis":     [_safe_stat(stats.kurtosis, rendements[c]) for c in rendements.columns],
            "Min (%)":      (rendements.min() * 100).round(3).values,
            "Max (%)":      (rendements.max() * 100).round(3).values,
        }, index=rendements.columns)
        stats_df.index.name = "Ticker"
        st.dataframe(stats_df, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : OPTIMISATION MARKOWITZ  (NOUVEAU)
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "📐 Optimisation":
    st.title("Optimisation de Portefeuille — Markowitz")

    if st.session_state["rendements"] is None:
        st.info("💡 Commencez par charger un portefeuille dans la page **Portefeuille**.")
        st.stop()

    rendements = st.session_state["rendements"]

    st.markdown("""
    <div class='info-box'>
    <b style='color:var(--amber,#F0B429)'>Théorie Moderne du Portefeuille (Markowitz, 1952)</b> — La frontière
    efficiente représente l'ensemble des portefeuilles offrant le <b>meilleur rendement pour
    un niveau de risque donné</b>. Trois allocations optimales sont proposées : maximisation
    du ratio de Sharpe, minimisation de la variance, et équipondération.
    </div>""", unsafe_allow_html=True)

    col_rf, col_btn, _ = st.columns([1,1,2])
    with col_rf:
        rf = st.number_input("Taux sans risque (%/an)", min_value=0.0, max_value=10.0,
                              value=3.0, step=0.1) / 100
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        opt_btn = st.button("▶  Optimiser le portefeuille", type="primary", use_container_width=True)

    if opt_btn:
        with st.spinner("Calcul de la frontière efficiente…"):
            opt_results = optimiser_portefeuille(rendements, rf)
            st.session_state["opt_results"] = opt_results
        st.success("✅ Optimisation terminée.")

    if st.session_state["opt_results"]:
        opt = st.session_state["opt_results"]

        st.markdown("<div class='section-header'>Frontière Efficiente</div>", unsafe_allow_html=True)
        fig_fe = fig_frontiere_efficiente(opt)
        st.pyplot(fig_fe, use_container_width=True); plt.close(fig_fe)

        st.markdown("<div class='section-header'>Allocations optimales</div>", unsafe_allow_html=True)
        cols = st.columns(3)
        styles_ = [("sharpe","#C9A84C"), ("minvar","#1db87a"), ("equi","#a855f7")]
        for col, (key, color) in zip(cols, styles_):
            p = opt[key]
            r, v, s = p["stats"]
            with col:
                st.markdown(f"""
                <div class='markowitz-card'>
                  <div style='font-size:11px;color:{color};font-family:DM Mono,monospace;
                       text-transform:uppercase;letter-spacing:1px;margin-bottom:8px'>{p['label']}</div>
                  <div style='font-size:13px;color:var(--txt-primary,#E8EDF5);line-height:2;font-family:DM Mono,monospace'>
                  📈 Rdt ann. : <b style='color:{color}'>{r*100:.2f}%</b><br>
                  📊 Vol. ann. : <b style='color:var(--txt-primary,#E8EDF5)'>{v*100:.2f}%</b><br>
                  ⭐ Sharpe   : <b style='color:{color}'>{s:.3f}</b>
                  </div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-header'>Répartition des poids</div>", unsafe_allow_html=True)
        fig_w = fig_poids_portefeuilles(opt)
        st.pyplot(fig_w, use_container_width=True); plt.close(fig_w)

        st.markdown("<div class='section-header'>Tableau des poids détaillés</div>", unsafe_allow_html=True)
        df_poids = pd.DataFrame({
            "Actif": opt["tickers"],
            "Sharpe Max. (%)":   [f"{w*100:.1f}%" for w in opt["sharpe"]["weights"]],
            "Variance Min. (%)": [f"{w*100:.1f}%" for w in opt["minvar"]["weights"]],
            "Équipondéré (%)":   [f"{w*100:.1f}%" for w in opt["equi"]["weights"]],
        }).set_index("Actif")
        st.dataframe(df_poids, use_container_width=True)

        with st.expander("ℹ️ Hypothèses & Limites du modèle"):
            st.markdown("""
            - **Positions longues uniquement** (contrainte w ≥ 0)
            - **Rendements supposés normaux** — en pratique, les rendements présentent des queues épaisses
            - **Paramètres historiques** — la frontière est sensible à la période d'estimation
            - **Pas de coûts de transaction** — les poids théoriques ignorent les frais de rééquilibrage
            - **Recommandation** : combiner l'allocation Sharpe Max avec la VaR TVE-GARCH pour un pilotage complet du risque
            """)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : VAR
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "📉 Calcul VaR":
    st.title("Calcul de la Value at Risk")

    if st.session_state["rendements"] is None:
        st.info("💡 Commencez par charger un portefeuille dans la page **Portefeuille**.")
        st.stop()

    rendements = st.session_state["rendements"]
    pv         = st.session_state["pv"] or 10_000_000
    port_r     = rendements.mean(axis=1).dropna().values

    st.markdown("<div class='section-header'>Paramètres de calcul</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1: horizon = st.slider("Horizon (jours)", 1, 10, 1)
    with c2:
        conf_options = st.multiselect("Niveaux de confiance",
                                       [0.90, 0.95, 0.975, 0.99],
                                       default=[0.95, 0.99],
                                       format_func=lambda x: f"{x*100:.1f}%")
    with c3:
        methodes_sel = st.multiselect("Méthodes à calculer",
                                       ["Historique","Variance-Covariance","RiskMetrics",
                                        "Cornish-Fisher","GARCH(1,1)","TVE (POT)","TVE-GARCH"],
                                       default=["Historique","Variance-Covariance",
                                                "RiskMetrics","Cornish-Fisher",
                                                "GARCH(1,1)","TVE (POT)","TVE-GARCH"])

    btn2, _ = st.columns([1,3])
    with btn2:
        calc_btn = st.button("▶  Calculer les 7 VaR", type="primary", use_container_width=True)

    if calc_btn:
        if not conf_options:
            st.warning("Sélectionnez au moins un niveau de confiance.")
        else:
            with st.spinner("Calcul en cours…"):
                engine = VaREngine(port_r, pv, horizon)
                var_results = engine.compute_all(tuple(sorted(conf_options)))
                var_results = {k:v for k,v in var_results.items() if k in methodes_sel}
                st.session_state["var_results"] = var_results
            st.success(f"✅ VaR calculée — {len(var_results)} méthodes × {len(conf_options)} niveaux.")

    if st.session_state["var_results"]:
        var_results = st.session_state["var_results"]
        alphas_used = sorted(list(list(var_results.values())[0].keys()))
        alpha_display = st.select_slider("Afficher pour :",
                                          options=alphas_used,
                                          format_func=lambda x: f"{x*100:.0f}%",
                                          value=alphas_used[-1])

        st.markdown("<div class='section-header'>Résultats par méthode</div>", unsafe_allow_html=True)
        METHODE_RECOMMANDEE = "TVE-GARCH"
        cols = st.columns(min(len(var_results), 4))
        for i, (method, res) in enumerate(var_results.items()):
            r = res.get(alpha_display, {})
            var_val = r.get("VaR", np.nan)
            pct_val = r.get("VaR_pct", np.nan)
            es_val  = r.get("ES",  np.nan)
            # Protection NaN pour affichage
            var_val = 0.0 if not np.isfinite(var_val) else var_val
            pct_val = 0.0 if not np.isfinite(pct_val) else pct_val
            es_val  = 0.0 if not np.isfinite(es_val)  else es_val
            rec = f'<span class="badge-rec">★ Recommandé</span>' if method == METHODE_RECOMMANDEE else ""
            with cols[i % len(cols)]:
                st.markdown(f"""
                <div class='var-card'>
                  <div class='var-card-title'>{method}{rec}</div>
                  <div class='var-card-value'>{var_val/1000:.1f} k€</div>
                  <div class='var-card-pct'>VaR {alpha_display*100:.0f}% · {pct_val*100:.3f}%</div>
                  <div class='var-card-es'>ES : {es_val/1000:.1f} k€</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-header'>Tableau comparatif</div>", unsafe_allow_html=True)
        rows = []
        for m, res in var_results.items():
            row = {"Méthode": m}
            for a in alphas_used:
                r = res.get(a, {})
                row[f"VaR {a*100:.0f}% (€)"]  = f"{r.get('VaR',0):,.0f}"
                row[f"VaR {a*100:.0f}% (%)"]  = f"{r.get('VaR_pct',0)*100:.3f}%"
                row[f"ES {a*100:.0f}% (€)"]   = f"{r.get('ES',0):,.0f}"
            row["Paramètres"] = list(var_results[m].values())[0].get("params","")
            rows.append(row)
        df_var = pd.DataFrame(rows).set_index("Méthode")
        st.dataframe(df_var, use_container_width=True)

        st.markdown("<div class='section-header'>Graphique comparatif</div>", unsafe_allow_html=True)
        fig_v = fig_var_comparaison(var_results, alpha_display, pv)
        st.pyplot(fig_v, use_container_width=True); plt.close(fig_v)

        st.markdown("<div class='section-header'>Distribution des rendements</div>", unsafe_allow_html=True)
        fig_d = fig_distribution(port_r, var_results)
        st.pyplot(fig_d, use_container_width=True); plt.close(fig_d)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : BACKTESTING
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "🧪 Backtesting":
    st.title("Backtesting des modèles de VaR")

    if st.session_state["var_results"] is None:
        st.info("💡 Calculez d'abord les VaR dans la page **Calcul VaR**.")
        st.stop()

    rendements   = st.session_state["rendements"]
    var_results  = st.session_state["var_results"]
    port_r       = rendements.mean(axis=1).dropna().values
    alphas_used  = sorted(list(list(var_results.values())[0].keys()))

    st.markdown("""
    <div class='info-box'>
    <b style='color:var(--amber,#F0B429)'>Test de Kupiec (POF)</b> — Vérifie si la fréquence d'exceptions
    est statistiquement conforme au niveau de confiance déclaré. <b>LR ~ χ²(1).</b><br><br>
    <b style='color:var(--amber,#F0B429)'>Test de Christoffersen (CC)</b> — Teste l'indépendance temporelle
    des exceptions. Un clustering signale un modèle insensible aux chocs.<br><br>
    <span style='color:var(--signal-green,#00C896)'>✅ p-value &gt; 5%</span> → modèle validé &nbsp;
    <span style='color:var(--risk-red,#FF4D6D)'>❌ p-value ≤ 5%</span> → modèle rejeté
    </div>""", unsafe_allow_html=True)

    btn3, _ = st.columns([1,3])
    with btn3:
        bt_btn = st.button("▶  Lancer le backtesting", type="primary", use_container_width=True)

    if bt_btn:
        with st.spinner("Backtesting en cours…"):
            bt_results = {}
            for method, res in var_results.items():
                bt_results[method] = {}
                for a in alphas_used:
                    var_pct = res.get(a, {}).get("VaR_pct", np.nan)
                    if np.isnan(var_pct): continue
                    k  = kupiec_test(port_r, var_pct, a)
                    cc = christoffersen_test(port_r, var_pct)
                    bt_results[method][a] = {"kupiec": k, "cc": cc}
            st.session_state["bt_results"] = bt_results
        st.success("✅ Backtesting terminé.")

    if st.session_state["bt_results"]:
        bt_results = st.session_state["bt_results"]
        rows = []
        for m, alphas in bt_results.items():
            for a, res in alphas.items():
                k, cc = res["kupiec"], res["cc"]
                rows.append({
                    "Méthode": m, "CL": f"{a*100:.0f}%",
                    "Exceptions": k["N"], "T": k["T"],
                    "Taux obs.": f"{k['rate']*100:.2f}%",
                    "Taux att.": f"{(1-a)*100:.2f}%",
                    "Kupiec LR": round(k["LR"],3),
                    "Kupiec p":  round(k["p_value"],4),
                    "Kupiec ✓":  "✅ OK" if k["valid"] else "❌",
                    "CC p":      round(cc.get("p_value_ind",0),4),
                    "CC ✓":      "✅ OK" if cc.get("valid") else "❌",
                })
        st.dataframe(pd.DataFrame(rows).set_index("Méthode"), use_container_width=True)

        st.markdown("<div class='section-header'>Graphique p-values (Kupiec)</div>", unsafe_allow_html=True)
        fig_bt = fig_backtesting({m: {a: {"p_value": bt_results[m][a]["kupiec"]["p_value"]}
                                       for a in bt_results[m]} for m in bt_results})
        st.pyplot(fig_bt, use_container_width=True); plt.close(fig_bt)

        st.markdown("<div class='section-header'>Verdict synthétique</div>", unsafe_allow_html=True)
        alpha_v = alphas_used[-1]
        cols_v = st.columns(len(bt_results))
        for col, (m, alphas) in zip(cols_v, bt_results.items()):
            res = alphas.get(alpha_v, {})
            k, cc = res.get("kupiec",{}), res.get("cc",{})
            k_ok, cc_ok = k.get("valid",False), cc.get("valid",False)
            score = "✅ Validé" if k_ok and cc_ok else ("⚠️ Partiel" if k_ok or cc_ok else "❌ Rejeté")
            color = "#1db87a" if k_ok and cc_ok else ("#C9A84C" if k_ok or cc_ok else "#e05252")
            with col:
                st.markdown(f"""
                <div class='var-card' style='text-align:center'>
                  <div class='var-card-title'>{m}</div>
                  <div style='font-size:15px;font-weight:700;color:{color};margin:6px 0'>{score}</div>
                  <div style='font-size:10px;color:var(--txt-secondary,#7A8BA8);font-family:DM Mono,monospace'>
                  Kupiec p={k.get('p_value',0):.4f}<br>CC p={cc.get('p_value_ind',0):.4f}
                  </div>
                </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : STRESS-TESTING  (NOUVEAU)
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "🔥 Stress-Testing":
    st.title("Stress-Testing — Scénarios Historiques")

    if st.session_state["rendements"] is None:
        st.info("💡 Commencez par charger un portefeuille dans la page **Portefeuille**.")
        st.stop()

    rendements  = st.session_state["rendements"]
    var_results = st.session_state["var_results"]
    pv          = st.session_state["pv"] or 10_000_000
    port_r      = rendements.mean(axis=1).values

    if var_results is None:
        st.warning("⚠️ Les VaR ne sont pas encore calculées. Les comparaisons seront limitées.")
        var_results = {}

    st.markdown("""
    <div class='info-box'>
    Le <b>stress-testing</b> mesure l'impact de scénarios de marché extrêmes sur le portefeuille.
    Contrairement à la VaR (probabiliste), les stress-tests appliquent des <b>chocs déterministes</b>
    calibrés sur des crises historiques réelles. Exigé par <b>Bâle III/IV</b> et les superviseurs (EBA, BCE).
    </div>""", unsafe_allow_html=True)

    st.markdown("<div class='section-header'>Sélection des scénarios</div>", unsafe_allow_html=True)
    scenarios_choisis = st.multiselect(
        "Scénarios à analyser",
        list(SCENARIOS_STRESS.keys()),
        default=list(SCENARIOS_STRESS.keys()),
    )

    c_alpha, c_btn, _ = st.columns([1,1,2])
    with c_alpha:
        alpha_stress = st.selectbox("Niveau VaR de référence", [0.95, 0.99],
                                     format_func=lambda x: f"{x*100:.0f}%", index=1)
    with c_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        stress_btn = st.button("▶  Lancer le stress-test", type="primary", use_container_width=True)

    if stress_btn and scenarios_choisis:
        with st.spinner("Calcul des scénarios de stress…"):
            stress_results = {}
            for sc_name in scenarios_choisis:
                sc = SCENARIOS_STRESS[sc_name]
                stress_results[sc_name] = appliquer_stress(
                    port_r, pv, sc, var_results, alpha=alpha_stress
                )
            st.session_state["stress_results"] = stress_results
        st.success(f"✅ {len(stress_results)} scénario(s) calculé(s).")

    if st.session_state["stress_results"]:
        stress_results = st.session_state["stress_results"]

        st.markdown("<div class='section-header'>Résultats par scénario</div>", unsafe_allow_html=True)
        cols_per_row = 3
        sc_list = list(stress_results.items())
        for row_start in range(0, len(sc_list), cols_per_row):
            cols = st.columns(cols_per_row)
            for col, (sc_name, sr) in zip(cols, sc_list[row_start:row_start+cols_per_row]):
                sc_info = SCENARIOS_STRESS.get(sc_name, {})
                ratio = sr["ratio"]
                ratio_txt = f"×{ratio:.2f}" if not np.isnan(ratio) else "—"
                alert = "🔴" if (not np.isnan(ratio) and ratio > 2) else ("🟡" if (not np.isnan(ratio) and ratio > 1.5) else "🟢")
                with col:
                    st.markdown(f"""
                    <div class='stress-card'>
                      <div class='stress-title'>{alert} {sc_name.split("(")[0].strip()}</div>
                      <div style='font-size:10px;color:var(--txt-secondary,#7A8BA8);margin-bottom:8px'>{sc_info.get('date','')}</div>
                      <div class='stress-val'>{sr['pnl_stress']/1000:+.0f} k€</div>
                      <div style='font-size:11px;color:var(--txt-secondary,#7A8BA8);margin-top:4px;font-family:DM Mono,monospace'>
                        P&L stressé<br>
                        VaR stressée : <b style='color:#e87373'>{sr['var_stress']/1000:.0f} k€</b><br>
                        Ratio : <b style='color:var(--amber,#F0B429)'>{ratio_txt}</b> vs VaR normale<br>
                        Choc marché : <b style='color:var(--risk-red,#FF4D6D)'>{sr['choc']*100:.2f}%</b>
                      </div>
                    </div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-header'>Analyse comparative</div>", unsafe_allow_html=True)
        fig_st = fig_stress_comparaison(stress_results, pv)
        st.pyplot(fig_st, use_container_width=True); plt.close(fig_st)

        st.markdown("<div class='section-header'>Tableau de synthèse</div>", unsafe_allow_html=True)
        rows_st = []
        for sc_name, sr in stress_results.items():
            sc_info = SCENARIOS_STRESS.get(sc_name, {})
            ratio = sr["ratio"]
            alert = "🔴 ALERTE" if (not np.isnan(ratio) and ratio > 2) else (
                    "🟡 VIGILANCE" if (not np.isnan(ratio) and ratio > 1.5) else "🟢 OK")
            rows_st.append({
                "Scénario":          sc_name.split("(")[0].strip(),
                "Date":              sc_info.get("date",""),
                "Choc marché (%)":   f"{sr['choc']*100:.2f}%",
                "P&L stressé (k€)":  f"{sr['pnl_stress']/1000:+.1f}",
                "VaR stressée (k€)": f"{sr['var_stress']/1000:.1f}",
                "VaR normale (k€)":  f"{sr['var_normal']/1000:.1f}",
                "Ratio ×":           f"{ratio:.2f}" if not np.isnan(ratio) else "—",
                "Statut":            alert,
            })
        st.dataframe(pd.DataFrame(rows_st).set_index("Scénario"), use_container_width=True)

        with st.expander("ℹ️ Méthodologie des stress-tests"):
            st.markdown("""
            **Choc de marché** : rendement journalier appliqué au portefeuille équipondéré.

            **VaR stressée** : recalculée avec la volatilité multipliée par le facteur de choc du scénario,
            selon la méthode Variance-Covariance : VaR_stress = z_{α} × σ_stress × PV

            **Ratio** : VaR_stressée / VaR_normale → mesure l'amplification du risque en période de crise.

            **Seuils d'alerte** :
            - 🟢 Ratio < 1.5 → Impact limité
            - 🟡 1.5 ≤ Ratio < 2.0 → Vigilance renforcée
            - 🔴 Ratio ≥ 2.0 → Alerte — révision des limites recommandée
            """)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : REPORTING
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "📊 Reporting":
    st.title("Génération des Rapports")

    if st.session_state["var_results"] is None:
        st.info("💡 Calculez d'abord les VaR dans la page **Calcul VaR**.")
        st.stop()

    rendements  = st.session_state["rendements"]
    var_results = st.session_state["var_results"]
    bt_results  = st.session_state["bt_results"] or {}
    pv          = st.session_state["pv"] or 10_000_000
    opt_results = st.session_state.get("opt_results")
    port_r      = rendements.mean(axis=1).values
    alphas_used = sorted(list(list(var_results.values())[0].keys()))

    st.markdown("""
    <div class='info-box'>
    Générez automatiquement un <b>rapport Excel</b> de suivi opérationnel (4 feuilles)
    et un <b>rapport PDF</b> exécutif avec graphiques. Conformes aux exigences Bâle III/IV.
    </div>""", unsafe_allow_html=True)

    st.markdown("<div class='section-header'>Aperçu des résultats</div>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    alpha_99 = max(alphas_used)
    best = "TVE-GARCH" if "TVE-GARCH" in var_results else list(var_results.keys())[-1]
    var_best = var_results[best][alpha_99]["VaR"]
    es_best  = var_results[best][alpha_99]["ES"]
    n_ok = sum(1 for m in bt_results for a in bt_results[m]
               if bt_results[m][a]["kupiec"]["valid"]) if bt_results else 0
    n_total = sum(len(bt_results[m]) for m in bt_results) if bt_results else 0

    c1.metric("Méthode recommandée", best)
    c2.metric(f"VaR 99% ({best})", f"{var_best/1000:.1f} k€")
    c3.metric(f"ES 99% ({best})",  f"{es_best/1000:.1f} k€")
    c4.metric("Backtests validés",  f"{n_ok}/{n_total}" if n_total else "—")

    st.markdown("<div class='section-header'>Téléchargements</div>", unsafe_allow_html=True)
    col_xlsx, col_pdf = st.columns(2)

    bt_fmt = {}
    for m, alphas in bt_results.items():
        bt_fmt[m] = {}
        for a, res in alphas.items():
            bt_fmt[m][a] = res

    with col_xlsx:
        st.markdown("""
        <div class='var-card'>
          <div class='var-card-title'>📊 Rapport Excel</div>
          <div style='font-size:12px;color:var(--txt-primary,#E8EDF5);margin:8px 0;line-height:1.6'>
          4 feuilles : Résumé VaR · Backtesting · Markowitz · Données<br>
          Formatage professionnel, tableaux structurés, prêt pour la Direction.
          </div>
        </div>""", unsafe_allow_html=True)
        if HAS_XLSX:
            with st.spinner("Génération Excel…"):
                xlsx_bytes = generer_excel(rendements, var_results, bt_fmt, pv, opt_results)
            if xlsx_bytes:
                st.download_button("⬇  Télécharger le rapport Excel",
                                   data=xlsx_bytes, file_name="VaR_Report_v4.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
        else:
            st.warning("openpyxl non installé : pip install openpyxl")

    with col_pdf:
        st.markdown("""
        <div class='var-card'>
          <div class='var-card-title'>📄 Rapport PDF</div>
          <div style='font-size:12px;color:var(--txt-primary,#E8EDF5);margin:8px 0;line-height:1.6'>
          Couverture · VaR · Backtesting · Graphiques intégrés<br>
          Mise en page exécutive, prêt à imprimer / envoyer.
          </div>
        </div>""", unsafe_allow_html=True)
        if HAS_PDF:
            with st.spinner("Génération PDF…"):
                fv  = fig_var_comparaison(var_results, alpha_99, pv)
                fd  = fig_distribution(port_r, var_results)
                pdf_bytes = generer_pdf(var_results, bt_fmt, {}, pv, fv, fd)
            if pdf_bytes:
                st.download_button("⬇  Télécharger le rapport PDF",
                                   data=pdf_bytes, file_name="VaR_Risk_Report_v4.pdf",
                                   mime="application/pdf",
                                   use_container_width=True)
        else:
            st.warning("reportlab non installé : pip install reportlab")

    with st.expander("📦 Déploiement Streamlit Cloud"):
        st.markdown("**`requirements.txt`** à placer à la racine du dépôt GitHub :")
        st.code("""streamlit>=1.32
yfinance>=0.2.36
pandas>=2.0
numpy>=1.26
scipy>=1.12
matplotlib>=3.8
openpyxl>=3.1
reportlab>=4.1""", language="text")
        st.markdown("""
        **Étapes :**
        1. Pousser `app.py` + `requirements.txt` sur GitHub
        2. Aller sur [share.streamlit.io](https://share.streamlit.io)
        3. **New app** → sélectionner le repo → `app.py` → **Deploy**
        """)
